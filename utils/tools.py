# -*- coding: utf-8 -*-
"""
@Time    : 2022/8/23 13:12
@Author  : itlubber
@Site    : itlubber.art
"""

import os
import six
import toad
import joblib
import warnings
import numpy as np
import pandas as pd
from tqdm import tqdm
import scorecardpy as sc
from datetime import datetime
import matplotlib.pyplot as plt
from optbinning import OptimalBinning
from sklearn.metrics import make_scorer
from sklearn.model_selection import train_test_split

from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font


def init_setting(font_path="./utils/matplot_chinese.ttf"):
    import warnings
    import matplotlib
    from matplotlib import font_manager
    warnings.filterwarnings("ignore")
    pd.options.display.float_format = '{:.4f}'.format
    pd.set_option('display.max_colwidth', 300)
    plt.style.use('seaborn-ticks')
    matplotlib.font_manager.fontManager.addfont(font_path)
    matplotlib.rcParams['font.family'] = font_manager.FontProperties(fname=font_path).get_name()
    matplotlib.rcParams['axes.unicode_minus'] = False
    
    
# warnings.filterwarnings("ignore")
# pd.set_option('display.width', 5000)
# plt.rcParams["font.sans-serif"]=["SimHei"] #设置字体
# plt.rcParams["axes.unicode_minus"]=False #该语句解决图像中的“-”负号的乱码问题


try:
    feature_describe = pd.read_excel("变量字典及字段解释.xlsx", sheet_name="数据字段表", header=0, engine="openpyxl", usecols=[0, 1])
    feature_describe = feature_describe.drop_duplicates(subset=["变量名称"], keep="last")
    feature_dict = dict(zip(feature_describe["变量名称"], feature_describe["含义"]))
except:
    feature_dict = {}


def ks_score(y, y_pred):
    return toad.KS(y_pred[:, 1], y)


ks_score = make_scorer(ks_score, needs_proba=True)


def round_float(num):
    if ~pd.isnull(num) and isinstance(num, float):
        return float(str(num).split(".")[0] + "." + str(num).split(".")[1][:4])
    else:
        return num


def feature_bins(bins):
    if isinstance(bins, list): bins = np.array(bins)
    EMPTYBINS = len(bins) if not isinstance(bins[0], (set, list, np.ndarray)) else -1

    l = []
    if np.issubdtype(bins.dtype, np.number):
        has_empty = len(bins) > 0 and np.isnan(bins[-1])
        if has_empty: bins = bins[:-1]
        sp_l = ["负无穷"] + [round_float(b) for b in bins.tolist()] + ["正无穷"]
        for i in range(len(sp_l) - 1): l.append('['+str(sp_l[i])+' , '+str(sp_l[i+1])+')')
        if has_empty: l.append('缺失值')
    else:
        for keys in bins:
            keys_update = set()
            for key in keys:
                if pd.isnull(key) or key == "nan":
                    keys_update.add("缺失值")
                elif key.strip() == "":
                    keys_update.add("空字符串")
                else:
                    keys_update.add(key)
            label = ','.join(keys_update)
            l.append(label)

    return {i if b != "缺失值" else EMPTYBINS: b for i, b in enumerate(l)}


def feature_bin_stats(data, feature, combiner=None, target="target", rules={}, empty_separate=True, method='cart', min_samples=0.15, max_n_bins=3, gamma=0.01, monotonic_trend="auto_asc_desc", feature_dict={}):
    # if combiner is None:
    #     combiner = toad.transform.Combiner()
    #     combiner.fit(data[[feature, target]], target, empty_separate=empty_separate, method=method, min_samples=min_samples) 
    if feature not in rules:
        if data[feature].nunique(dropna=True) < 3:
            splits = []
            for v in data[feature].unique():
                if not pd.isnull(v):
                    splits.append(v)

            if str(data[feature].dtypes) in ["object", "string", "category"]:
                rule = {feature: [[s] for s in splits]}
                rule[feature].append([[np.nan]])
            else:
                rule = {feature: sorted(splits) + [np.nan]}
        else:
            try:
                y = data[target]
                if str(data[feature].dtypes) in ["object", "string", "category"]:
                    dtype = "categorical"
                    x = data[feature].astype("category").values
                else:
                    dtype = "numerical"
                    x = data[feature].values
                _combiner = OptimalBinning(feature, dtype=dtype, max_n_bins=max_n_bins, monotonic_trend=monotonic_trend, gamma=gamma).fit(x, y)
                if _combiner.status == "OPTIMAL":
                    rule = {feature: [s.tolist() if isinstance(s, np.ndarray) else s for s in _combiner.splits] + [[np.nan] if dtype == "categorical" else np.nan]}
                else:
                    raise "OptimalBinning error"
            except Exception as e:
                if method not in ["dt", "chi", ]:
                    method = "chi"
                _combiner = toad.transform.Combiner()
                _combiner.fit(data[[feature, target]], target, empty_separate=empty_separate, method=method, min_samples=min_samples)
                rule = _combiner.export()

    if combiner is None:
        combiner = toad.transform.Combiner()
    
    combiner.update(rule)
    
    if rules and isinstance(rules, list): rules = {feature: rules}
    if rules and isinstance(rules, dict): combiner.update(rules)

    # feature_bin = combiner.export()[feature]
    # feature_bin_dict = format_bins(np.array(feature_bin))
    
    df_bin = combiner.transform(data[[feature, target]], labels=False)
    
    table = df_bin[[feature, target]].groupby([feature, target]).agg(len).unstack()
    table.columns.name = None
    table = table.rename(columns = {0 : '好样本数', 1 : '坏样本数'}).fillna(0)
    table["指标名称"] = feature
    table["指标含义"] = feature_dict.get(feature, "")
    table = table.reset_index().rename(columns={feature: "分箱"})
    # table["分箱"] = table["分箱"].map(feature_bin_dict)

    table['样本总数'] = table['好样本数'] + table['坏样本数']
    table['样本占比'] = table['样本总数'] / table['样本总数'].sum()
    table['好样本占比'] = table['好样本数'] / table['好样本数'].sum()
    table['坏样本占比'] = table['坏样本数'] / table['坏样本数'].sum()
    table['坏样本率'] = table['坏样本数'] / table['样本总数']
    
    table = table.fillna(0.)
    
    table['分档WOE值'] = table.apply(lambda x : np.log(x['坏样本占比'] / (x['好样本占比'] + 1e-6)),axis=1)
    table['分档IV值'] = table.apply(lambda x : (x['坏样本占比'] - x['好样本占比']) * np.log(x['坏样本占比'] / (x['好样本占比'] + 1e-6)), axis=1)
    table['指标IV值'] = table['分档IV值'].sum()
    
    table["LIFT值"] = table['坏样本率'] / (table["坏样本数"].sum() / table["样本总数"].sum())
    table["累积LIFT值"] = table["LIFT值"].cumsum()
    
    return table[['指标名称', "指标含义", '分箱', '样本总数', '样本占比', '好样本数', '好样本占比', '坏样本数', '坏样本占比', '坏样本率', '分档WOE值', '分档IV值', '指标IV值', 'LIFT值', '累积LIFT值']]


def plot_bin(binx, title="", show_iv=True, show_na=True, colors=["#2639E9", "#a29bfe", "#ff7675"], figsize=(10, 8)):
    if not show_na:
        binx = binx[binx["分箱"] != "缺失值"].reset_index(drop=True)
    # y_right_max
    y_right_max = np.ceil(binx['坏样本率'].max()*10)
    if y_right_max % 2 == 1: y_right_max=y_right_max+1
    if y_right_max - binx['坏样本率'].max()*10 <= 0.3: y_right_max = y_right_max+2
    y_right_max = y_right_max/10
    if y_right_max>1 or y_right_max<=0 or y_right_max is np.nan or y_right_max is None: y_right_max=1
    ## y_left_max
    y_left_max = np.ceil(binx['样本占比'].max()*10)/10
    if y_left_max>1 or y_left_max<=0 or y_left_max is np.nan or y_left_max is None: y_left_max=1
    # title
    title_string = binx.loc[0,'指标名称']+"  (iv:"+str(round(binx['分档IV值'].sum(),4))+")" if show_iv else binx.loc[0,'指标名称']
    title_string = title + '-' + title_string if title else title_string
    # param
    ind = np.arange(len(binx.index))    # the x locations for the groups
    width = 0.35       # the width of the bars: can also be len(x) sequence
    ###### plot ###### 
    fig, ax1 = plt.subplots(figsize=figsize)
    ax2 = ax1.twinx()
    # ax1
    p1 = ax1.bar(ind, binx['好样本占比'], width, color=colors[1])
    p2 = ax1.bar(ind, binx['坏样本占比'], width, bottom=binx['好样本占比'], color=colors[2])
    for i in ind:
        ax1.text(i, binx.loc[i,'样本占比']*1.02, str(round(binx.loc[i,'样本占比']*100,1))+'%, '+str(binx.loc[i,'样本总数']), ha='center')
    # ax2
    ax2.plot(ind, binx['坏样本率'], marker='o', color=colors[0])
    for i in ind:
        ax2.text(i, binx.loc[i,'坏样本率']*1.02, str(round(binx.loc[i,'坏样本率']*100,1))+'%', color=colors[0], ha='center')
    # settings
    ax1.set_ylabel('样本分布情况')
    ax2.set_ylabel('坏样本率', color=colors[0])
    ax1.set_yticks(np.arange(0, y_left_max+0.2, 0.2))
    ax2.set_yticks(np.arange(0, y_right_max+0.2, 0.2))
    ax2.tick_params(axis='y', colors=colors[0])
    plt.xticks(ind, binx['分箱'], fontsize=12)
    plt.title(title_string, loc='center')
    plt.legend((p2[0], p1[0]), ('好样本', '坏样本'), loc='upper right')
    

# def bin_plot(feature_table, feature="", desc="", figsize=(8, 6), colors=['#8E8BFE', '#FEA3A2', '#9394E7'], max_len=35, save=None):
#     feature_table = feature_table.copy()
# 
#     feature_table["分箱"] = feature_table["分箱"].apply(lambda x: x if re.match("^\[.*\)$", x) else str(x)[:max_len] + "..")
# 
#     # 绘制好坏样本分布情况
#     fig, ax1 = plt.subplots(figsize=figsize)
#     ax1.barh(feature_table['分箱'], feature_table['好样本数'], color=colors[0], label='好样本')
#     ax1.barh(feature_table['分箱'], feature_table['坏样本数'], left=feature_table['好样本数'], color=colors[1], label='坏样本')
#     ax1.set_xlabel('样本数')
# 
#     # 绘制坏样本率的分布情况
#     ax2 = ax1.twiny()
#     ax2.plot(feature_table['坏样本率'], feature_table['分箱'], colors[2], label='坏样本率', linestyle='-.')
#     ax2.set_xlabel('坏样本率: 坏样本数 / 样本总数')
# 
#     for i, rate in enumerate(feature_table['坏样本率']):
#         ax2.scatter(rate, i, color=colors[2], s=3)
# 
#     # 在图像对应位置显示样本总数和坏样本率
#     for i, v in feature_table[['样本总数', '好样本数', '坏样本数', '坏样本率', '样本占比']].iterrows():
#         ax1.text(v['样本总数'] / 2, i + len(feature_table) / 60, f"{int(v['好样本数'])}:{int(v['坏样本数'])}:{v['样本占比']:.1%}:{v['坏样本率']:.1%}")
# 
#     # 逆转y轴顺序
#     ax1.invert_yaxis()
#     
#     desc = desc if desc else feature
# 
#     # 添加一个标题
#     fig.suptitle(f'变量 {desc} 分箱图\n\n')
# 
#     # 合并图例
#     handles1, labels1 = ax1.get_legend_handles_labels()
#     handles2, labels2 = ax2.get_legend_handles_labels()
#     fig.legend(handles1 + handles2, labels1 + labels2, loc='upper center', ncol=len(labels1 + labels2), bbox_to_anchor=(0.5, 0.95), frameon=False)
# 
#     # 调整布局，使分箱信息能够完全显示
#     plt.tight_layout()
# 
#     if save:
#         if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
#             os.makedirs(os.path.dirname(save))
# 
#         fig.savefig(save, dpi=240, format="png", bbox_inches="tight")
    
    
def cal_psi(train, test, feature, combiner=None):
    # feature_bin = combiner.export()[feature]
    # feature_bin_dict = format_bins(np.array(feature_bin))
    try:
        A = (combiner.transform(train[[feature]]).value_counts() / len(train[[feature]])).reset_index().rename(columns={feature: "分箱", 0: "A"})
        E = (combiner.transform(test[[feature]]).value_counts() / len(test[[feature]])).reset_index().rename(columns={feature: "分箱", 0: "E"})
    except:
        A = (combiner.transform(train[[feature]])[feature].value_counts() / len(train)).reset_index().rename(columns={"index": "分箱", feature: "A"})
        E = (combiner.transform(test[[feature]])[feature].value_counts() / len(test)).reset_index().rename(columns={"index": "分箱", feature: "E"})
    df_psi = A.merge(E, on="分箱", how="outer").fillna(0.)
    # df_psi["分箱"] = df_psi["分箱"].map(feature_bin_dict)
    df_psi["分档PSI"] = (df_psi["A"] - df_psi["E"]) * np.log(df_psi["A"] / (df_psi["E"] + 1e-6))
    df_psi["指标PSI"] = df_psi["分档PSI"].replace(np.inf, 0).sum()
    
    return df_psi[["分箱", "分档PSI", "指标PSI"]]


def itlubber_border(border, color):
    if len(border) == 3:
        return Border(
            left=Side(border_style=border[0], color=color[0]),
            right=Side(border_style=border[1], color=color[1]),
            bottom=Side(border_style=border[2], color=color[2]),
        )
    else:
        return Border(
            left=Side(border_style=border[0], color=color[0]),
            right=Side(border_style=border[1], color=color[1]),
            bottom=Side(border_style=border[2], color=color[2]),
            top=Side(border_style=border[3], color=color[3]),
        )


def render_excel(excel_name, sheet_name=None, conditional_columns=[], freeze=None, merge_rows=[], percent_columns=[], theme_color="2639E9", conditional_color="9980FA", font="楷体", fontsize=10, max_column_width=50, header=True, start_row=0, n_jobs=4, bar=True, border=True):
    workbook = load_workbook(excel_name)
    
    if sheet_name and isinstance(sheet_name, str):
        sheet_names = [sheet_name]
    else:
        sheet_names = workbook.get_sheet_names()
    
    merge_rows = [i + start_row if header else i + start_row - 1 for i in merge_rows]
    
    for sheet_name in sheet_names:
        worksheet = workbook.get_sheet_by_name(sheet_name)
        
        def add_conditional_formatting(column, theme_color="FDA7DF"):
            worksheet.conditional_formatting.add(f'{column}2:{column}{worksheet.max_row}', DataBarRule(start_type='min', end_type='max', color=theme_color))
        
        for conditional_column in conditional_columns:
            add_conditional_formatting(f"{conditional_column}", theme_color=conditional_color)
        
        def render_cell(row_index, row):
            if row_index > start_row:
                if header and row_index == start_row + 1:
                    for col_index, cell in enumerate(row, start=1):
                        cell.font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
                        cell.fill = PatternFill(fill_type="solid", start_color=theme_color)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        
                        if col_index == 1:
                            cell.border = itlubber_border(["medium", "thin", "medium", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
                        elif col_index == len(row):
                            cell.border = itlubber_border(["thin", "medium", "medium", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])
                        else:
                            cell.border = itlubber_border(["thin", "thin", "medium", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
                else:
                    for col_index, cell in enumerate(row, start=1):
                        cell.font = Font(size=fontsize, name=font, color="000000")
                        cell.fill = PatternFill(fill_type="solid", start_color="FFFFFF")
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        
                        if col_index in percent_columns:
                            # cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
                            cell.number_format = "0.00%"
                        else:
                            pass
                            # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        
                        if row_index == worksheet.max_row:
                            if col_index == 1:
                                cell.border = itlubber_border(["medium", "thin", "medium"], [theme_color, "FFFFFF", theme_color])
                            elif col_index == len(row):
                                cell.border = itlubber_border(["thin", "medium", "medium"], ["FFFFFF", theme_color, theme_color])
                            else:
                                cell.border = itlubber_border(["thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color])
                        else:
                            if merge_rows in [[], None] or (row_index - 1 in merge_rows):
                                if col_index == 1:
                                    cell.border = itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", theme_color])
                                elif col_index == len(row):
                                    cell.border = itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, theme_color])
                                else:
                                    cell.border = itlubber_border(["thin", "thin", "thin"], ["FFFFFF", "FFFFFF", theme_color])
                            else:
                                if col_index == 1:
                                    cell.border = itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", "FFFFFF"])
                                elif col_index == len(row):
                                    cell.border = itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, "FFFFFF"])
                                else:
                                    cell.border = itlubber_border(["thin", "thin", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        
        if border:
            iterrows = tqdm(enumerate(worksheet.rows, start=1), total=worksheet.max_row - 1) if bar else enumerate(worksheet.rows, start=1)
            if n_jobs > 0:
                joblib.Parallel(n_jobs=n_jobs)(joblib.delayed(render_cell)(row_index, row) for row_index, row in iterrows)
            else:
                for row_index, row in iterrows:
                    render_cell(row_index, row)
        
        feature_table = pd.read_excel(
            excel_name, sheet_name=sheet_name, engine="openpyxl"
        )
        feature_table_len_max = feature_table.apply(lambda x: [(len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i)) for i in x]).max()
        for i in feature_table.columns:
            # 列的字母
            j = list(feature_table.columns)
            column_letter = [chr(j.index(i) + 65) if j.index(i) <= 25 else 'A' + chr(j.index(i) - 26 + 65) ][0]
            # 列的宽度
            columns_length = (len(str(i).encode('utf-8')) - len(str(i)))/2 + len(str(i))
            data_max_length = feature_table_len_max[i]
            column_width = [data_max_length if columns_length < data_max_length else columns_length][0]
            column_width = [column_width if column_width <= max_column_width else max_column_width][0] + 3
            # 更改列的宽度
            worksheet.column_dimensions['{}'.format(column_letter)].width = column_width
            
        if freeze:
            worksheet.freeze_panes = freeze
    
    workbook.save(excel_name)
    workbook.close()
    

def run_feature_table(feature, train=None, feature_dict=None, rules={}, combiner=None, target="target", return_feature=False):
    table = feature_bin_stats(train, feature, feature_dict=feature_dict, rules=rules, combiner=combiner)
    df_psi = cal_psi(train[[feature, target]], test[[feature, target]], feature, combiner=combiner)
    
    table = table.merge(df_psi, on="分箱", how="left")
    
    feature_bin = combiner.export()[feature]
    feature_bin_dict = format_bins(np.array(feature_bin))
    table["分箱"] = table["分箱"].map(feature_bin_dict)
    
    if return_feature:
        return feature, table
    else:
        return table
    
    
def render_dataframe(df, row_height=0.4, font_size=14,
                     header_color='#2639E9', row_colors=['#dae3f3', 'w'], edge_color='w',
                     bbox=[0, 0, 1, 1], header_columns=0,
                     ax=None, save=None, **kwargs):
    data = df.copy()
    for col in data.select_dtypes('datetime'):
        data[col] = data[col].dt.strftime("%Y-%m-%d")

    for col in data.select_dtypes('float'):
        data[col] = data[col].apply(lambda x: np.nan if pd.isnull(x) else round(x, 4))

    cols_width = [max(data[col].apply(lambda x:len(str(x).encode())).max(), len(str(col).encode())) / 8. for col in data.columns]

    if ax is None:
        size = (sum(cols_width), (len(data) + 1) * row_height)
        fig, ax = plt.subplots(figsize=size)
        ax.axis('off')

    mpl_table = ax.table(cellText=data.values, colWidths=cols_width, bbox=bbox, colLabels=data.columns, **kwargs)

    mpl_table.auto_set_font_size(False)
    mpl_table.set_fontsize(font_size)

    for k, cell in  six.iteritems(mpl_table._cells):
        cell.set_edgecolor(edge_color)
        if k[0] == 0 or k[1] < header_columns:
            cell.set_text_props(weight='bold', color='w')
            cell.set_facecolor(header_color)
        else:
            cell.set_facecolor(row_colors[k[0]%len(row_colors)])

    if save:
        if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
            os.makedirs(os.path.dirname(save))

        fig.savefig(save, dpi=240, format="png", bbox_inches="tight")

    return fig


if __name__ == '__main__':
    from functools import partial
    from multiprocessing import Pool
    data = sc.germancredit()
    
    # 测试数据
    data["target"] = data["creditability"].replace({'good':0,'bad':1})
    data["credit.amount"].loc[0] = np.nan
    data["status.of.existing.checking.account"].loc[0] = np.nan
    data["test_a"] = 0.
    data["test_a"].loc[0] = np.nan
    data["test_b"] = ""
    data["test_b"].loc[0] = np.nan
    data["test_c"] = np.nan
    
    # data = data.replace("", np.nan)
    
    train, test = train_test_split(data, test_size=0.3,)
    
    target = "target"
    cols = ["test_a", "test_b", "test_c", "status.of.existing.checking.account", "credit.amount"]
    
    combiner = toad.transform.Combiner()
    # combiner.fit(data[cols + [target]], target, empty_separate=True, method="chi", min_samples=0.2)
    
    # 保存结果至 EXCEL 文件
    output_excel_name = f"指标有效性验证-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    output_sheet_name = "指标有效性"
    tables = {}
    merge_row_number = []
    
    # _run_feature_table = partial(run_feature_table, train=train, feature_dict=feature_dict, rules={}, combiner=combiner, target=target, return_feature=True)
    # all_feature_tables = joblib.Parallel(n_jobs=4)(joblib.delayed(_run_feature_table)(feature) for feature in cols)
    
    # for feature, table in all_feature_tables:
    #     merge_row_number.append(len(table))
    #     tables[feature] = table
    
    for feature in cols:
        table = feature_bin_stats(train, feature, feature_dict=feature_dict, rules={}, combiner=combiner)
        print(train.shape)
        df_psi = cal_psi(train[[feature, target]], test[[feature, target]], feature, combiner=combiner)
        
        table = table.merge(df_psi, on="分箱", how="left")
        
        feature_bin = combiner.export()[feature]
        feature_bin_dict = format_bins(np.array(feature_bin))
        table["分箱"] = table["分箱"].map(feature_bin_dict)
    
        table = run_feature_table(feature)
        # plot_bin(table, show_na=True)
        merge_row_number.append(len(table))
        tables[feature] = table

    merge_row_number = np.cumsum(merge_row_number).tolist()
    feature_table = pd.concat(tables, ignore_index=True).round(6)
    feature_table["分档WOE值"] = feature_table["分档WOE值"].fillna(np.inf)
    
    workbook = load_workbook(output_excel_name) if os.path.exists(output_excel_name) else None
    writer = pd.ExcelWriter(output_excel_name, engine="openpyxl")
    
    if workbook:
        writer.book = workbook
        writer.sheets = {ws.title: ws for ws in workbook.worksheets}
        start_row = writer.book.get_sheet_by_name(output_sheet_name).max_row 
    else:
        start_row = 0
        
    feature_table.to_excel(writer, sheet_name=output_sheet_name, index=False, header=True, startcol=0, startrow=start_row)
    
    writer.close()
    
    render_excel(output_excel_name, sheet_name=output_sheet_name, conditional_columns=["J", "N"], freeze="D2", merge_rows=merge_row_number, percent_columns=[5, 7, 9, 10], start_row=start_row, header=False if start_row > 0 else True)
    # render_excel("变量字典及字段解释.xlsx")
    combiner.export(to_json=f"rules_{datetime.now().strftime('%Y-%m-%d')}.json")
