# -*- coding: utf-8 -*-
"""
@Time    : 2023/2/15 17:55
@Author  : itlubber
@Site    : itlubber.art
"""
import math
import sys
import re
import matplotlib
import matplotlib.font_manager as font_manager
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string


from model import *
from utils.excel_writer import ExcelWriter


plt.style.use('seaborn-ticks')
# plt.style.use('seaborn-white')
# plt.rcParams.update({'font.size': 14})


def pyplot_chinese(font_path='utils/matplot_chinese.ttf'):
    # matplotlib.rcParams['font.size'] = 20
    matplotlib.font_manager.fontManager.addfont(font_path)
    matplotlib.rcParams['font.family'] = font_manager.FontProperties(fname=font_path).get_name()
    matplotlib.rcParams['axes.unicode_minus']=False


pyplot_chinese(font_path='utils/杨任东竹石体-Medium.ttf')


target = "creditability"
data = sc.germancredit()
data[target] = data[target].map({"good": 0, "bad": 1})

train, test = train_test_split(data, test_size=0.3, shuffle=True, stratify=data[target])
oot = data.copy()

feature_pipeline = Pipeline([
    ("preprocessing_select", FeatureSelection(target=target, engine="scorecardpy")),
    ("combiner", Combiner(target=target, min_samples=0.2)),
    ("transform", WOETransformer(target=target)),
    # ("processing_select", FeatureSelection(target=target, engine="scorecardpy")),
    ("stepwise", StepwiseSelection(target=target)),
])

feature_pipeline.fit(train)

woe_train = feature_pipeline.transform(train)
woe_test = feature_pipeline.transform(test)
woe_oot = feature_pipeline.transform(oot)

# # save all bin_plot
# _combiner = feature_pipeline.named_steps["combiner"]
# for col in woe_train.columns:
#     if col != target:
#         _combiner.bin_plot(train, col, labels=True, save=f"model_report/bin_plots/train_{col}.png")
#         _combiner.bin_plot(test, col, labels=True, save=f"model_report/bin_plots/test_{col}.png")
#         _combiner.bin_plot(oot, col, labels=True, save=f"model_report/bin_plots/oot_{col}.png")

# logistic = StatsLogisticRegression(target=target)
logistic = ITLubberLogisticRegression(target=target)

logistic.fit(woe_train)

y_pred_train = logistic.predict_proba(woe_train.drop(columns=target))[:, 1]
y_pred_test = logistic.predict_proba(woe_test.drop(columns=target))[:, 1]
y_pred_oot = logistic.predict_proba(woe_oot.drop(columns=target))[:, 1]

ScoreCard.ks_plot(y_pred_train, train[target], save="model_report/lr_ksplot_train.png", figsize=(10, 5))
ScoreCard.ks_plot(y_pred_test, test[target], save="model_report/lr_ksplot_test.png", figsize=(10, 5))
ScoreCard.ks_plot(y_pred_oot, oot[target], save="model_report/lr_ksplot_oot.png", figsize=(10, 5))

summary = logistic.summary().reset_index().rename(columns={"index": "Features"})

train_corr = logistic.corr(woe_train, save="model_report/train_corr.png")
test_corr = logistic.corr(woe_test, save="model_report/test_corr.png")
oot_corr = logistic.corr(woe_oot, save="model_report/oot_corr.png")

train_report = logistic.report(woe_train)
test_report = logistic.report(woe_test)
oot_report = logistic.report(woe_oot)

print("train: ", toad.metrics.KS(y_pred_train, train[target]), toad.metrics.AUC(y_pred_train, train[target]))
print("test: ", toad.metrics.KS(y_pred_test, test[target]), toad.metrics.AUC(y_pred_test, test[target]))
print("oot: ", toad.metrics.KS(y_pred_oot, oot[target]), toad.metrics.AUC(y_pred_oot, oot[target]))


card = ScoreCard(target=target, pipeline=feature_pipeline, pretrain_lr=logistic)
card.fit(woe_train)

train["score"] = card.predict(train)
test["score"] = card.predict(test)
oot["score"] = card.predict(oot)


def sample_distribution(df, date="date", target="target", user_count="count", save="model_report/sample_time_count.png", figsize=(10, 6)):
    temp = df.groupby([df[date].dt.strftime("%Y-%m"), df[target].map({0: "好样本", 1: "坏样本"})])[user_count].sum().unstack()
    fig, ax1 = plt.subplots(1, 1, figsize=figsize)
    temp.plot(kind='bar', stacked=True, ax=ax1, color=["#8E8BFE", "#FEA3A2"], hatch="/", legend=False)
    ax1.tick_params(axis='x', labelrotation=-90)
    ax1.set(xlabel=None)
    ax1.set_ylabel('样本数')
    ax1.set_title('不同时点数据集样本分布情况\n\n')

    ax2 = plt.twinx()
    (temp["坏样本"] / temp.sum(axis=1)).plot(ax=ax2, color="#9394E7", marker=".", linewidth=2, label="坏样本率")
    # sns.despine()

    # 合并图例
    handles1, labels1 = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(handles1 + handles2, labels1 + labels2, loc='upper center', ncol=len(labels1 + labels2), bbox_to_anchor=(0.5, 0.94), frameon=False)
    # ax1.legend(frameon=False, labels=["good", "bad"], loc='upper right')
    # ax2.legend(loc='upper left', frameon=False, labels=["bad rate"])

    plt.tight_layout()

    if save:
        if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
            os.makedirs(os.path.dirname(save))

        fig.savefig(save, dpi=240, format="png", bbox_inches="tight")

    temp = temp.reset_index().rename(columns={"date": "日期", 0: "好样本", 1: "坏样本"})
    temp["样本总数"] = temp["坏样本"] + temp["好样本"]
    temp["样本占比"] = temp["样本总数"] / temp["样本总数"].sum()
    temp["好样本占比"] = temp["好样本"] / temp["好样本"].sum()
    temp["坏样本占比"] = temp["坏样本"] / temp["坏样本"].sum()
    temp["坏样本率"] = temp["坏样本"] / temp["样本总数"]

    return temp[["日期", "样本总数", "样本占比", "好样本", "好样本占比", "坏样本", "坏样本占比", "坏样本率"]]


def bin_plot(feature_table, feature="", figsize=(15, 8), colors=['#8E8BFE', '#FEA3A2', '#9394E7'], max_len=35, save=None):
    feature_table = feature_table.copy()

    feature_table["分箱"] = feature_table["分箱"].apply(lambda x: x if re.match("^\[.*\)$", x) else str(x)[:max_len] + "..")

    # 绘制好坏样本分布情况
    fig, ax1 = plt.subplots(figsize=figsize)
    ax1.barh(feature_table['分箱'], feature_table['好样本数'], color=colors[0], label='好样本', hatch="/")
    ax1.barh(feature_table['分箱'], feature_table['坏样本数'], left=feature_table['好样本数'], color=colors[1], label='坏样本', hatch="\\")
    ax1.set_xlabel('样本数')

    # 绘制坏样本率的分布情况
    ax2 = ax1.twiny()
    ax2.plot(feature_table['坏样本率'], feature_table['分箱'], colors[2], label='坏样本率', linestyle='-.')
    ax2.set_xlabel('坏样本率: 坏样本数 / 样本总数')

    for i, rate in enumerate(feature_table['坏样本率']):
        ax2.scatter(rate, i, color=colors[2])

    # 在图像对应位置显示样本总数和坏样本率
    for i, v in feature_table[['样本总数', '好样本数', '坏样本数', '坏样本率']].iterrows():
        ax1.text(v['样本总数'] / 2, i + len(feature_table) / 60, f"{int(v['好样本数'])}:{int(v['坏样本数'])}:{v['坏样本率']:.2%}")

    # 逆转y轴顺序
    ax1.invert_yaxis()

    # 添加一个标题
    fig.suptitle(f'变量 {feature} 分箱图\n\n')

    # 合并图例
    handles1, labels1 = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    fig.legend(handles1 + handles2, labels1 + labels2, loc='upper center', ncol=len(labels1 + labels2), bbox_to_anchor=(0.5, 0.925), frameon=False)

    # 调整布局，使分箱信息能够完全显示
    plt.tight_layout()

    if save:
        if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
            os.makedirs(os.path.dirname(save))

        fig.savefig(save, dpi=240, format="png", bbox_inches="tight")


writer = ExcelWriter(style_excel="./utils/报告输出模版.xlsx", theme_color="8E8BFE")


# ////////////////////////////////////// 样本说明 ///////////////////////////////////// #
df = pd.DataFrame({
    "date": pd.date_range(start="2021-01-01", end="2022-06-30"),
    "target": np.random.randint(0, 2, 546),
    "count": np.random.randint(0, 100, 546),
})

total_count = len(data)
dataset_summary = pd.DataFrame(
    [
        ["建模样本", "2022-01-01", "2023-01-31", len(data), len(data) / total_count, data[target].sum(), data[target].sum() / len(data), ""],
        ["训练集", "2022-01-01", "2023-12-31", len(train), len(train) / total_count, train[target].sum(), train[target].sum() / len(train), ""],
        ["测试集", "2022-01-01", "2023-12-31", len(test), len(test) / total_count, test[target].sum(), test[target].sum() / len(test), ""],
        ["跨时间验证集", "2023-01-01", "2023-01-31", len(oot), len(oot) / total_count, oot[target].sum(), oot[target].sum() / len(oot), ""],
    ],
    columns=["数据集", "开始时间", "结束时间", "样本总数", "样本占比", "坏客户数", "坏客户占比", "备注"],
)

worksheet = writer.get_sheet_by_name("汇总信息")

# 样本总体分布情况
start_row, start_col = 2, 2
end_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value="样本总体分布情况", style="header")
end_row, end_col = writer.insert_df2sheet(worksheet, dataset_summary, (end_row + 1, start_col), header=True)

writer.set_number_format(worksheet, f"{get_column_letter(end_col - 2)}{end_row - len(dataset_summary)}:{get_column_letter(end_col - 2)}{end_row}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(end_col - 4)}{end_row - len(dataset_summary)}:{get_column_letter(end_col - 4)}{end_row}", "0.00%")

# 建模样本时间分布情况
temp = sample_distribution(df, date="date", target="target", user_count="count", save="model_report/all_sample_time_count.png")
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="建模样本时间分布情况", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/all_sample_time_count.png", (end_row, start_col), figsize=(720, 370))
end_row, end_col = writer.insert_df2sheet(worksheet, temp.T.reset_index(), (end_row, start_col), header=False)

writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 1}:{get_column_letter(end_col)}{end_row - 1}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 2}:{get_column_letter(end_col)}{end_row - 2}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 4}:{get_column_letter(end_col)}{end_row - 4}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 6}:{get_column_letter(end_col)}{end_row - 6}", "0.00%")

# 训练集样本时间分布情况
temp = sample_distribution(df, date="date", target="target", user_count="count", save="model_report/train_sample_time_count.png")
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="训练集样本时间分布情况", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/train_sample_time_count.png", (end_row, start_col), figsize=(720, 370))
end_row, end_col = writer.insert_df2sheet(worksheet, temp.T.reset_index(), (end_row, start_col), header=False)

writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 1}:{get_column_letter(end_col)}{end_row - 1}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 2}:{get_column_letter(end_col)}{end_row - 2}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 4}:{get_column_letter(end_col)}{end_row - 4}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 6}:{get_column_letter(end_col)}{end_row - 6}", "0.00%")

# 测试集样本时间分布情况
temp = sample_distribution(df, date="date", target="target", user_count="count", save="model_report/test_sample_time_count.png")
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="测试集样本时间分布情况", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/test_sample_time_count.png", (end_row, start_col), figsize=(720, 370))
end_row, end_col = writer.insert_df2sheet(worksheet, temp.T.reset_index(), (end_row, start_col), header=False)

writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 1}:{get_column_letter(end_col)}{end_row - 1}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 2}:{get_column_letter(end_col)}{end_row - 2}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 4}:{get_column_letter(end_col)}{end_row - 4}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 6}:{get_column_letter(end_col)}{end_row - 6}", "0.00%")

# 跨时间验证集样本时间分布情况
temp = sample_distribution(df, date="date", target="target", user_count="count", save="model_report/oot_sample_time_count.png")
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="跨时间验证集样本时间分布情况", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/oot_sample_time_count.png", (end_row, start_col), figsize=(720, 370))
end_row, end_col = writer.insert_df2sheet(worksheet, temp.T.reset_index(), (end_row, start_col), header=False)

writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 1}:{get_column_letter(end_col)}{end_row - 1}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 2}:{get_column_letter(end_col)}{end_row - 2}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 4}:{get_column_letter(end_col)}{end_row - 4}", "0.00%")
writer.set_number_format(worksheet, f"{get_column_letter(start_col)}{end_row - 6}:{get_column_letter(end_col)}{end_row - 6}", "0.00%")


# ////////////////////////////////////// 模型报告 ///////////////////////////////////// #

# 逻辑回归拟合情况
worksheet = writer.get_sheet_by_name("逻辑回归拟合结果")
start_row, start_col = 2, 2

end_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value="逻辑回归拟合效果", style="header")
# worksheet.merge_cells(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + len(summary.columns) - 1)}{start_row}")
# worksheet[f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + len(summary.columns) - 1)}{start_row}"].style = "header"
logistic.plot_weights(save="model_report/logistic_train.png")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/logistic_train.png", (end_row + 2, start_col))
end_row, end_col = writer.insert_df2sheet(worksheet, summary, (end_row + 1, start_col))

conditional_column = get_column_letter(start_col + summary.columns.get_loc("Coef."))
writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(summary)}', f'{conditional_column}{end_row}')

# worksheet.merge_cells(f"{get_column_letter(start_col)}{end_row + 2}:{get_column_letter(start_col + len(train_report.columns) - 1)}{end_row + 2}")
# worksheet[f"{get_column_letter(start_col)}{end_row + 2}"].style = "header"
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="训练数据集拟合报告", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/lr_ksplot_train.png", (end_row, start_col), figsize=(480, 270))
end_row, end_col = writer.insert_df2sheet(worksheet, train_report, (end_row + 1, start_col))

# worksheet.merge_cells(f"{get_column_letter(start_col)}{end_row + 2}:{get_column_letter(start_col + len(test_report.columns) - 1)}{end_row + 2}")
# worksheet[f"{get_column_letter(start_col)}{end_row + 2}"].style = "header"
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="测试数据集拟合报告", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/lr_ksplot_test.png", (end_row, start_col), figsize=(480, 270))
end_row, end_col = writer.insert_df2sheet(worksheet, test_report, (end_row + 1, start_col))

# worksheet.merge_cells(f"{get_column_letter(start_col)}{end_row + 2}:{get_column_letter(start_col + len(oot_report.columns) - 1)}{end_row + 2}")
# worksheet[f"{get_column_letter(start_col)}{end_row + 2}"].style = "header"
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="跨时间验证集拟合报告", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/lr_ksplot_oot.png", (end_row, start_col), figsize=(480, 270))
end_row, end_col = writer.insert_df2sheet(worksheet, oot_report, (end_row + 1, start_col))


# ////////////////////////////////////// 特征概述 ///////////////////////////////////// #

# 模型变量概览
feature_describe = pd.DataFrame([
    ["status_account", "支票账户状态"], ["duration", "借款周期"], ["credit_histor", "历史信用"], ["purpose", "借款目的"], ["amount", "信用额度"], ["svaing_account", "储蓄账户状态"], ["present_emp", "当前就业状态"], ["income_rate", "分期付款占可支配收入百分比"], ["personal_status", "性别与婚姻状态"], ["other_debtors", "他人担保信息"], ["residence_info", "现居住地"], ["property", "财产状态"], ["age", "年龄"], ["inst_plans", "其他分期情况"], ["housing", "房产状态"], ["num_credits", "信用卡数量"], ["job", "工作状态"], ["dependents", "赡养人数"], ["telephone", "电话号码注册情况"], ["foreign_worke", "是否有海外工作经历"],
], columns=["变量名称", "变量含义"])

worksheet = writer.get_sheet_by_name("模型变量信息")
start_row, start_col = 2, 2
end_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value="入模变量信息", style="header")
end_row, end_col = writer.insert_df2sheet(worksheet, feature_describe.reset_index().rename(columns={"index": "序号"}), (end_row + 1, start_col))

# 变量分布情况
data_info = toad.detect(data[card.rules.keys()]).reset_index().rename(columns={"index": "变量名称", "type": "变量类型", "size": "样本个数", "missing": "缺失值", "unique": "唯一值个数"})
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="变量分布情况", style="header")
end_row, end_col = writer.insert_df2sheet(worksheet, data_info, (end_row + 1, start_col))

# 变量相关性
data_corr = logistic.corr(feature_pipeline.transform(train), save="model_report/data_corr.png", annot=False)
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="变量相关性", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/data_corr.png", (end_row + 1, start_col), figsize=(700, 500))
end_row, end_col = writer.insert_df2sheet(worksheet, data_corr.reset_index().rename(columns={"index": ""}), (end_row + 1, start_col))

conditional_column = f"{get_column_letter(start_col + 1)}{end_row - len(data_corr)}:{get_column_letter(end_col - 1)}{end_row - 1}"
worksheet.conditional_formatting.add(conditional_column, ColorScaleRule(start_type='num', start_value=-1.0, start_color='8E8BFE', mid_type='num', mid_value=0., mid_color='FFFFFF', end_type='num', end_value=1.0, end_color='8E8BFE'))


# 变量分箱信息
_combiner = feature_pipeline.named_steps["combiner"]

end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="变量分箱信息", style="header")
for col in card.rules.keys():
    feature_table = card.feature_bin_stats(data, col, target=target, desc="逻辑回归入模变量", combiner=card.combiner)
    # _combiner.bin_plot(data, col, labels=True, save=f"model_report/bin_plots/data_{col}.png")
    bin_plot(feature_table, feature=col, save=f"model_report/bin_plots/data_{col}.png")
    end_row, end_col = writer.insert_pic2sheet(worksheet, f"model_report/bin_plots/data_{col}.png", (end_row + 1, start_col), figsize=(700, 400))
    end_row, end_col = writer.insert_df2sheet(worksheet, feature_table, (end_row, start_col))

    for c in ["坏样本率", "LIFT值"]:
        conditional_column = get_column_letter(start_col + feature_table.columns.get_loc(c))
        writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(feature_table)}', f'{conditional_column}{end_row}')
        # conditional_column = get_column_letter(start_col + feature_table.columns.get_loc("LIFT值"))
        # writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(feature_table)}', f'{conditional_column}{end_row}')

    for c in ["样本占比", "好样本占比", "坏样本占比", "坏样本率", "LIFT值", "累积LIFT值"]:
        conditional_column = get_column_letter(start_col + feature_table.columns.get_loc(c))
        writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(feature_table)}:{conditional_column}{end_row}", "0.00%")


# ////////////////////////////////////// 评分卡说明 ///////////////////////////////////// #

# 评分卡刻度
scorecard_kedu = pd.DataFrame(
    [
        ["base_odds", card.base_odds, "根据业务经验设置的基础比率（违约概率/正常概率），估算方法：（1-样本坏客户占比）/坏客户占比"],
        ["base_score", card.base_score, "基础ODDS对应的分数"],
        ["rate", card.rate, "设置分数的倍率"],
        ["pdo", card.pdo, "表示分数增长PDO时，ODDS值增长到RATE倍"],
        ["B", card.offset, "补偿值，计算方式：pdo / ln(rate)"],
        ["A", card.factor, "刻度，计算方式：base_score - B * ln(base_odds)"],
    ],
    columns=["刻度项", "刻度值", "备注"],
)

worksheet = writer.get_sheet_by_name("评分卡结果")
start_row, start_col = 2, 2
end_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value="评分卡刻度", style="header")
end_row, end_col = writer.insert_df2sheet(worksheet, scorecard_kedu, (end_row + 1, start_col))

# 评分卡对应分数
card_points = card.export(to_frame=True).rename(columns={"name": "变量名称", "value": "变量分箱", "score": "对应分数"})
end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="评分卡分数", style="header")
end_row, end_col = writer.insert_df2sheet(worksheet, card_points, (end_row + 1, start_col), merge_column="变量名称")

# 评分效果
clip = 50
clip_start = max(math.ceil(train["score"].min() / clip) * clip, math.ceil(train["score"].quantile(0.01) / clip) * clip)
clip_end = min(math.ceil(train["score"].max() / clip) * clip, math.ceil(train["score"].quantile(0.99) / clip) * clip)
score_clip = [i for i in range(clip_start, clip_end, clip)]

train_score_rank = card.feature_bin_stats(train, "score", target=target, rules=score_clip, verbose=0, method="step", ks=True)
test_score_rank = card.feature_bin_stats(test, "score", target=target, rules=score_clip, verbose=0, method="step", ks=True)
oot_score_rank = card.feature_bin_stats(oot, "score", target=target, rules=score_clip, verbose=0, method="step", ks=True)

card.ks_plot(train["score"], train[target], title="Train Dataset", save="model_report/train_ksplot.png")
card.ks_plot(test["score"], test[target], title="Test Dataset", save="model_report/test_ksplot.png")
card.ks_plot(oot["score"], oot[target], title="OOT Dataset", save="model_report/oot_ksplot.png")

card.score_hist(train["score"], train[target], save="model_report/train_scorehist.png", bins=30, figsize=(13, 10))
card.score_hist(test["score"], test[target], save="model_report/test_scorehist.png", bins=30, figsize=(13, 10))
card.score_hist(oot["score"], oot[target], save="model_report/oot_scorehist.png", bins=30, figsize=(13, 10))


end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="训练数据集评分模型效果", style="header")
ks_row = end_row
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/train_ksplot.png", (ks_row, start_col))
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/train_scorehist.png", (ks_row, end_col))
end_row, end_col = writer.insert_df2sheet(worksheet, train_score_rank, (end_row + 1, start_col))

for c in ["坏样本率", "LIFT值", "分档KS值"]:
    conditional_column = get_column_letter(start_col + train_score_rank.columns.get_loc(c))
    writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(train_score_rank)}', f'{conditional_column}{end_row}')

for c in ["样本占比", "好样本占比", "坏样本占比", "坏样本率", "LIFT值", "累积LIFT值", "分档KS值"]:
    conditional_column = get_column_letter(start_col + train_score_rank.columns.get_loc(c))
    writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(train_score_rank)}:{conditional_column}{end_row}", "0.00%")

# conditional_column = get_column_letter(start_col + train_score_rank.columns.get_loc("坏样本率"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(train_score_rank)}', f'{conditional_column}{end_row}')
# conditional_column = get_column_letter(start_col + train_score_rank.columns.get_loc("LIFT值"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(train_score_rank)}', f'{conditional_column}{end_row}')
# conditional_column = get_column_letter(start_col + train_score_rank.columns.get_loc("分档KS值"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(train_score_rank)}', f'{conditional_column}{end_row}')


end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="测试数据集评分模型效果", style="header")
ks_row = end_row
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/test_ksplot.png", (ks_row, start_col))
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/test_scorehist.png", (ks_row, end_col))
end_row, end_col = writer.insert_df2sheet(worksheet, test_score_rank, (end_row + 1, start_col))

for c in ["坏样本率", "LIFT值", "分档KS值"]:
    conditional_column = get_column_letter(start_col + test_score_rank.columns.get_loc(c))
    writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(test_score_rank)}', f'{conditional_column}{end_row}')

for c in ["样本占比", "好样本占比", "坏样本占比", "坏样本率", "LIFT值", "累积LIFT值", "分档KS值"]:
    conditional_column = get_column_letter(start_col + test_score_rank.columns.get_loc(c))
    writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(test_score_rank)}:{conditional_column}{end_row}", "0.00%")

# conditional_column = get_column_letter(start_col + test_score_rank.columns.get_loc("坏样本率"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(test_score_rank)}', f'{conditional_column}{end_row}')
# conditional_column = get_column_letter(start_col + test_score_rank.columns.get_loc("LIFT值"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(test_score_rank)}', f'{conditional_column}{end_row}')
# conditional_column = get_column_letter(start_col + test_score_rank.columns.get_loc("分档KS值"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(test_score_rank)}', f'{conditional_column}{end_row}')


end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="跨时间验证集评分模型效果", style="header")
ks_row = end_row
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/oot_ksplot.png", (ks_row, start_col))
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/oot_scorehist.png", (ks_row, end_col))
end_row, end_col = writer.insert_df2sheet(worksheet, oot_score_rank, (end_row + 1, start_col))

for c in ["坏样本率", "LIFT值", "分档KS值"]:
    conditional_column = get_column_letter(start_col + oot_score_rank.columns.get_loc(c))
    writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row - len(oot_score_rank)}', f'{conditional_column}{end_row}')

for c in ["样本占比", "好样本占比", "坏样本占比", "坏样本率", "LIFT值", "累积LIFT值", "分档KS值"]:
    conditional_column = get_column_letter(start_col + oot_score_rank.columns.get_loc(c))
    writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(oot_score_rank)}:{conditional_column}{end_row}", "0.00%")

# conditional_column = get_column_letter(start_col + oot_score_rank.columns.get_loc("坏样本率"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(oot_score_rank)}', f'{conditional_column}{end_row}')
# conditional_column = get_column_letter(start_col + oot_score_rank.columns.get_loc("LIFT值"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(oot_score_rank)}', f'{conditional_column}{end_row}')
# conditional_column = get_column_letter(start_col + oot_score_rank.columns.get_loc("分档KS值"))
# writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(oot_score_rank)}', f'{conditional_column}{end_row}')


def score_psi(expected, actual, labels=["预期", "实际"], save=None, colors=['#8E8BFE', '#FEA3A2', '#9394E7'], figsize=(15, 8)):
    expected = expected.rename(columns={"分箱": "评分区间", "样本总数": f"{labels[0]}样本数", "样本占比": f"{labels[0]}样本占比", "坏样本率": f"{labels[0]}坏样本率"})
    actual = actual.rename(columns={"分箱": "评分区间", "样本总数": f"{labels[1]}样本数", "样本占比": f"{labels[1]}样本占比", "坏样本率": f"{labels[1]}坏样本率"})
    df_psi = expected.merge(actual, on="评分区间", how="outer").replace(np.nan, 0)
    df_psi[f"{labels[1]}% - {labels[0]}%"] = df_psi[f"{labels[1]}样本占比"] - df_psi[f"{labels[0]}样本占比"]
    df_psi[f"ln({labels[1]}% / {labels[0]}%)"] = np.log(df_psi[f"{labels[1]}样本占比"] / df_psi[f"{labels[0]}样本占比"])
    df_psi["分档PSI值"] = (df_psi[f"{labels[1]}% - {labels[0]}%"] * df_psi[f"ln({labels[1]}% / {labels[0]}%)"])
    df_psi = df_psi.fillna(0).replace(np.inf, 0).replace(-np.inf, 0)
    df_psi["总体PSI值"] = df_psi["分档PSI值"].sum()

    if save:
        if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
            os.makedirs(os.path.dirname(save))

        x = df_psi['评分区间']
        width = 0.35
        x_indexes = np.arange(len(x))
        fig, ax1 = plt.subplots(figsize=figsize)

        ax1.bar(x_indexes - width / 2, df_psi[f'{labels[0]}样本占比'], width, label=f'{labels[0]}样本占比', color=colors[0], hatch="/")
        ax1.bar(x_indexes + width / 2, df_psi[f'{labels[1]}样本占比'], width, label=f'{labels[1]}样本占比', color=colors[1], hatch="\\")

        ax1.set_ylabel('样本占比: 评分区间内样本数 / 样本总数')
        ax1.set_xticks(x_indexes)
        ax1.set_xticklabels(x)
        ax1.tick_params(axis='x', labelrotation=90)

        ax2 = ax1.twinx()
        ax2.plot(df_psi["评分区间"], df_psi[f"{labels[0]}坏样本率"], color=colors[0], label=f"{labels[0]}坏样本率", linestyle=(5, (10, 3)))
        ax2.plot(df_psi["评分区间"], df_psi[f"{labels[1]}坏样本率"], color=colors[1], label=f"{labels[1]}坏样本率", linestyle=(5, (10, 3)))

        ax2.scatter(df_psi["评分区间"], df_psi[f"{labels[0]}坏样本率"], marker=".")
        ax2.scatter(df_psi["评分区间"], df_psi[f"{labels[1]}坏样本率"], marker=".")

        ax2.set_ylabel('坏样本率: 坏样本数 / 样本总数')

        handles1, labels1 = ax1.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        fig.legend(handles1 + handles2, labels1 + labels2, loc='upper center', ncol=len(labels1 + labels2), bbox_to_anchor=(0.5, 0.94), frameon=False)

        fig.suptitle(f"{labels[0]} vs {labels[1]} 群体稳定性指数(PSI): {df_psi['分档PSI值'].sum():.4f}\n\n")

        fig.tight_layout()

        fig.savefig(save, dpi=240, format="png", bbox_inches="tight")

    return df_psi[["评分区间", f"{labels[0]}样本数", f"{labels[0]}样本占比", f"{labels[0]}坏样本率", f"{labels[1]}样本数", f"{labels[1]}样本占比", f"{labels[1]}坏样本率", f"{labels[1]}% - {labels[0]}%", f"ln({labels[1]}% / {labels[0]}%)", "分档PSI值", "总体PSI值"]]


train_test_score_psi = score_psi(train_score_rank, test_score_rank, labels=["训练数据集", "测试数据集"], save="model_report/train_test_psiplot.png")
train_oot_score_psi = score_psi(train_score_rank, oot_score_rank, labels=["训练数据集", "跨时间验证集"], save="model_report/train_oot_psiplot.png")
test_oot_score_psi = score_psi(test_score_rank, oot_score_rank, labels=["测试数据集", "跨时间验证集"], save="model_report/test_oot_psiplot.png")


end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="评分卡模型稳定性评估: 训练数据集 vs 测试数据集", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/train_test_psiplot.png", (end_row, start_col), figsize=(1000, 400))
end_row, end_col = writer.insert_df2sheet(worksheet, train_test_score_psi, (end_row + 1, start_col))

conditional_column = get_column_letter(start_col + train_test_score_psi.columns.get_loc("分档PSI值"))
writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(train_test_score_psi)}', f'{conditional_column}{end_row}')

for c in ["训练数据集样本占比", "训练数据集坏样本率", "测试数据集样本占比", "测试数据集坏样本率"]:
    conditional_column = get_column_letter(start_col + train_test_score_psi.columns.get_loc(c))
    writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(train_test_score_psi)}:{conditional_column}{end_row}", "0.00%")


end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="评分卡模型稳定性评估: 训练数据集 vs 跨时间验证集", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/train_oot_psiplot.png", (end_row, start_col), figsize=(1000, 400))
end_row, end_col = writer.insert_df2sheet(worksheet, train_oot_score_psi, (end_row + 1, start_col))

conditional_column = get_column_letter(start_col + train_oot_score_psi.columns.get_loc("分档PSI值"))
writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(train_oot_score_psi)}', f'{conditional_column}{end_row}')

for c in ["训练数据集样本占比", "训练数据集坏样本率", "跨时间验证集样本占比", "跨时间验证集坏样本率"]:
    conditional_column = get_column_letter(start_col + train_oot_score_psi.columns.get_loc(c))
    writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(train_oot_score_psi)}:{conditional_column}{end_row}", "0.00%")


end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="评分卡模型稳定性评估: 测试数据集 vs 跨时间验证集", style="header")
end_row, end_col = writer.insert_pic2sheet(worksheet, "model_report/test_oot_psiplot.png", (end_row, start_col), figsize=(1000, 400))
end_row, end_col = writer.insert_df2sheet(worksheet, test_oot_score_psi, (end_row + 1, start_col))

conditional_column = get_column_letter(start_col + test_oot_score_psi.columns.get_loc("分档PSI值"))
writer.add_conditional_formatting(worksheet, f'{conditional_column}{end_row-len(test_oot_score_psi)}', f'{conditional_column}{end_row}')

for c in ["跨时间验证集样本占比", "跨时间验证集坏样本率", "测试数据集样本占比", "测试数据集坏样本率"]:
    conditional_column = get_column_letter(start_col + test_oot_score_psi.columns.get_loc(c))
    writer.set_number_format(worksheet, f"{conditional_column}{end_row - len(test_oot_score_psi)}:{conditional_column}{end_row}", "0.00%")


# ////////////////////////////////////// 模型稳定性 ///////////////////////////////////// #
#
# worksheet = writer.get_sheet_by_name("模型稳定性")
# start_row, start_col = 2, 2
#
# # 变量 CSI 表
# end_row, end_col = writer.insert_value2sheet(worksheet, (start_row, start_col), value="入模变量稳定性指标 (Characteristic Stability Index, CSI)", style="header")
#
# # train vs test
#
# # 评分分布稳定性
# end_row, end_col = writer.insert_value2sheet(worksheet, (end_row + 2, start_col), value="模型评分稳定性指标 (Population Stability Index, PSI)", style="header")


writer.save("model_report/评分卡模型报告.xlsx")
