# -*- coding: utf-8 -*-
"""
@Time    : 2022/8/23 13:12
@Author  : itlubber
@Site    : itlubber.art
"""

import os
import toad
import warnings
import numpy as np
import pandas as pd
import scorecardpy as sc
from scorecardpy.perf import eva_pks, eva_proc
from optbinning import OptimalBinning
import matplotlib.pyplot as plt
from matplotlib import font_manager
import seaborn as sns
# import plotly.graph_objects as go
# from plotly.io import write_image
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

import scipy
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor

from sklearn.pipeline import Pipeline
from sklearn.metrics import roc_curve, auc
from sklearn.metrics import classification_report
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.utils.validation import check_is_fitted
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.base import BaseEstimator, TransformerMixin, ClassifierMixin

from processing import FeatureSelection, Combiner, WOETransformer, StepwiseSelection


warnings.filterwarnings("ignore")
pd.set_option('display.width', 5000)
# plt.rcParams["font.sans-serif"]=["SimHei"] #设置字体
# plt.rcParams["axes.unicode_minus"]=False #该语句解决图像中的“-”负号的乱码问题


def pyplot_chinese(font_path='utils/matplot_chinese.ttf'):
    font_manager.fontManager.addfont(font_path)
    plt.rcParams['font.family'] = font_manager.FontProperties(fname=font_path).get_name()
    plt.rcParams['axes.unicode_minus']=False


class StatsLogisticRegression(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", intercept=True):
        """
        基于statsmodels的逻辑回归方法

        Args:
            target: 数据集中标签名称，默认 target
            intercept: 是否包含截距，默认 True，即包含截距
        """
        self.intercept = intercept
        self.target = target
        self.classifier = None
        self.corr = None
        self.vif = None
        self.coef_normalization = None
        self.feature_names_ = None
        self.feature_importances_ = None
    
    def fit(self, x, y=None, vif=True, corr=True, normalization=True):
        self.feature_names_ = list(x.drop(columns=[self.target]).columns)
        self.feature_importances_ = self.feature_importances(x)
        
        if vif:
            self.vif = self.VIF(x)
            
        if normalization:
            _x = x.drop(columns=[self.target]).apply(lambda x: (x - np.mean(x)) / np.std(x))
            _y = x[self.target]
            lr_normalization = sm.Logit(_y, sm.add_constant(_x) if self.intercept else _x).fit()
            self.coef_normalization = pd.DataFrame(lr_normalization.params, columns=["coef_normalization"])
            
        if corr:
            self.corr = x.drop(columns=[self.target]).corr()
            
        if self.intercept:
            x = sm.add_constant(x)
        
        self.classes_ = x[self.target].unique()
        self.classifier = sm.Logit(x[self.target], x.drop(columns=[self.target])).fit()
        
        return self
    
    def transform(self, x):
        if self.intercept:
            x = sm.add_constant(x)
        
        return self.classifier.predict(x)
    
    def predict(self, x):
        return self.transform(x)
    
    def summary(self):
        describe = self.classifier.summary2()
        return describe
    
    def feature_importances(self, x):
        params = {
            "n_estimators": 256,
            "max_depth": 4,
            "min_samples_split": 5,
            "learning_rate": 1e-3,
            "loss": "deviance",
            "subsample": 0.9,
        }
        feature_importances_ = GradientBoostingClassifier(**params).fit(x.drop(columns=[self.target]), x[self.target]).feature_importances_
        return pd.DataFrame(feature_importances_, index=self.feature_names_, columns=["feature_importances"])
        
    def VIF(self, x):
        if self.intercept:
            x = sm.add_constant(x)
        
        x = x.drop(columns=[self.target])
        columns = x.columns
        vif = pd.DataFrame({"VIF": [variance_inflation_factor(np.matrix(x), i) for i in range(len(columns))]}, index=columns)
        
        return vif
    
    def WALD(self):
        return self.classifier.wald_test_terms().table[["statistic", "pvalue"]].rename(columns={"pvalue": "wald_test_pvalue", "statistic": "wald_test_statistic"})
    
    def report(self):
        return self.classifier.summary2().tables[1].join([self.coef_normalization, self.WALD(), self.vif, self.feature_importances_]), self.classifier.summary2().tables[0], self.corr
    
    def summary_save(self, excel_name="逻辑回归模型拟合效果.xlsx", sheet_name="逻辑回归拟合效果"):
        writer = pd.ExcelWriter(excel_name, engine='openpyxl')
        
        coef_report, summary_report, corr_report = self.report()
        summary_report.columns = ["逻辑回归模型拟合效果"] * summary_report.shape[1]
        summary_report.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startcol=0, startrow=2)
        coef_report.reset_index().rename(columns={"index": "variable"}).to_excel(writer, sheet_name=sheet_name, index=False, header=True, startcol=0, startrow=summary_report.shape[0] + 4)
        corr_report.to_excel(writer, sheet_name=sheet_name, index=True, header=True, startcol=0, startrow=summary_report.shape[0] + coef_report.shape[0] + 7)
        
        writer.save()
        writer.close()
        
        if os.path.exists(excel_name):
            workbook = load_workbook(excel_name)
            worksheet = workbook.get_sheet_by_name(sheet_name)
            worksheet["A1"].value = "逻辑回归模型报告"
            worksheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(f"A1:L1")
            
            workbook.save(excel_name)
            workbook.close()
        
        try:
            from processing import render_excel # From: https://github.com/itlubber/openpyxl-excel-style-template/blob/main/feature_bins.py
            render_excel(excel_name, sheet_name=sheet_name, max_column_width=25, merge_rows=np.cumsum([1, len(summary_report), 2, len(coef_report) + 1, 2, len(corr_report) + 1]).tolist())
        except:
            pass


class ITLubberLogisticRegression(LogisticRegression):
    """
    Extended Logistic Regression.
    Extends [sklearn.linear_model.LogisticRegression](https://scikit-learn.org/stable/modules/generated/sklearn.linear_model.LogisticRegression.html).
    This class provides the following extra statistics, calculated on `.fit()` and accessible via `.summary()`:
    - `cov_matrix_`: covariance matrix for the estimated parameters.
    - `std_err_intercept_`: estimated uncertainty for the intercept
    - `std_err_coef_`: estimated uncertainty for the coefficients
    - `z_intercept_`: estimated z-statistic for the intercept
    - `z_coef_`: estimated z-statistic for the coefficients
    - `p_value_intercept_`: estimated p-value for the intercept
    - `p_value_coef_`: estimated p-value for the coefficients
    
    Example:
    ```python
    feature_pipeline = Pipeline([
        ("preprocessing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("combiner", Combiner(target=target, min_samples=0.2)),
        ("transform", WOETransformer(target=target)),
        ("processing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("stepwise", StepwiseSelection(target=target)),
        # ("logistic", LogisticClassifier(target=target)),
        ("logistic", ITLubberLogisticRegression(target=target)),
    ])
    
    feature_pipeline.fit(train)
    summary = feature_pipeline.named_steps['logistic'].summary()
    ```
    
    An example output of `.summary()`:
    
    |                   |     Coef. |   Std.Err |        z |       P>|z| |    [ 0.025 |   0.975 ] |     VIF |
    |:------------------|----------:|----------:|---------:|------------:|-----------:|----------:|--------:|
    | const             | -0.844037 | 0.0965117 | -8.74544 | 2.22148e-18 | -1.0332    | -0.654874 | 1.05318 |
    | duration.in.month |  0.847445 | 0.248873  |  3.40513 | 0.000661323 |  0.359654  |  1.33524  | 1.14522 |
    """

    def __init__(self, target="target", penalty="l2", calculate_stats=True, dual=False, tol=0.0001, C=1.0, fit_intercept=True, intercept_scaling=1, class_weight=None, random_state=None, solver="lbfgs", max_iter=100, multi_class="auto", verbose=0, warm_start=False, n_jobs=None, l1_ratio=None,):
        """
        Extends [sklearn.linear_model.LogisticRegression.fit()](https://scikit-learn.org/stable/modules/generated/sklearn.linear_model.LogisticRegression.html).

        Args:
            target (str): your dataset's target name
            calculate_stats (bool): If true, calculate statistics like standard error during fit, accessible with .summary()
        """
        super().__init__(penalty=penalty, dual=dual, tol=tol, C=C, fit_intercept=fit_intercept, intercept_scaling=intercept_scaling, class_weight=class_weight, random_state=random_state, solver=solver, max_iter=max_iter, multi_class=multi_class, verbose=verbose, warm_start=warm_start, n_jobs=n_jobs, l1_ratio=l1_ratio,)
        self.target = target
        self.calculate_stats = calculate_stats

    def fit(self, x, sample_weight=None, **kwargs):
        y = x[self.target]
        x = x.drop(columns=[self.target])
        
        if not self.calculate_stats:
            return super().fit(x, y, sample_weight=sample_weight, **kwargs)

        x = self.convert_sparse_matrix(x)
        
        if isinstance(x, pd.DataFrame):
            self.names_ = ["const"] + [f for f in x.columns]
        else:
            self.names_ = ["const"] + [f"x{i}" for i in range(x.shape[1])]

        lr = super().fit(x, y, sample_weight=sample_weight, **kwargs)

        predProbs = self.predict_proba(x)

        # Design matrix -- add column of 1's at the beginning of your x matrix
        if lr.fit_intercept:
            x_design = np.hstack([np.ones((x.shape[0], 1)), x])
        else:
            x_design = x

        self.vif = [variance_inflation_factor(np.matrix(x_design), i) for i in range(x_design.shape[-1])]
        p = np.product(predProbs, axis=1)
        self.cov_matrix_ = np.linalg.inv((x_design * p[..., np.newaxis]).T @ x_design)
        std_err = np.sqrt(np.diag(self.cov_matrix_)).reshape(1, -1)

        # In case fit_intercept is set to True, then in the std_error array
        # Index 0 corresponds to the intercept, from index 1 onwards it relates to the coefficients
        # If fit intercept is False, then all the values are related to the coefficients
        if lr.fit_intercept:

            self.std_err_intercept_ = std_err[:, 0]
            self.std_err_coef_ = std_err[:, 1:][0]

            self.z_intercept_ = self.intercept_ / self.std_err_intercept_

            # Get p-values under the gaussian assumption
            self.p_val_intercept_ = scipy.stats.norm.sf(abs(self.z_intercept_)) * 2

        else:
            self.std_err_intercept_ = np.array([np.nan])
            self.std_err_coef_ = std_err[0]

            self.z_intercept_ = np.array([np.nan])

            # Get p-values under the gaussian assumption
            self.p_val_intercept_ = np.array([np.nan])

        self.z_coef_ = self.coef_ / self.std_err_coef_
        self.p_val_coef_ = scipy.stats.norm.sf(abs(self.z_coef_)) * 2

        return self
    
    def corr(self, data, save=None, annot=True):
        corr = data.drop(columns=[self.target]).corr()
        
        if save:
            self.corr_plot(data.drop(columns=[self.target]), save=save, annot=annot)
            
        return corr
    
    @staticmethod
    def corr_plot(data, figure_size=(16, 8),  fontsize=14, color=["#2639E9", "#F76E6C", "#FE7715"], mask=False, save=None, annot=True):
        corr = data.corr()
        corr_mask = np.zeros_like(corr, dtype = np.bool)
        corr_mask[np.triu_indices_from(corr_mask)] = True

        map_plot = toad.tadpole.tadpole.heatmap(
            corr,
            mask = corr_mask if mask else None,
            cmap = sns.diverging_palette(267, 267, n=10, s=100, l=40),
            vmax = 1,
            vmin = -1,
            center = 0,
            square = True,
            linewidths = .1,
            annot = annot,
            fmt = '.2f',
            figure_size = figure_size,
        )

        map_plot.tick_params(axis='x', labelrotation=270, labelsize=fontsize)
        map_plot.tick_params(axis='y', labelrotation=0, labelsize=fontsize)
        
        if save:
            if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
                os.makedirs(os.path.dirname(save))
            
            plt.savefig(save, dpi=240, format="png", bbox_inches="tight")
        
        return map_plot

    def report(self, data):
        report_dict = classification_report(data[self.target], self.predict(data.drop(columns=self.target)), output_dict=True, target_names=["好客户", "坏客户"])
        accuracy = report_dict.pop("accuracy")
        _report = pd.DataFrame(report_dict).T.reset_index().rename(columns={"index": "desc"})
        _report.loc[len(_report)] = ['accuracy', '', '', accuracy, len(data)]
        return _report

    def summary(self):
        """
        Puts the summary statistics of the fit() function into a pandas DataFrame.
        Returns:
            data (pandas DataFrame): The statistics dataframe, indexed by the column name
        """
        check_is_fitted(self)

        if not hasattr(self, "std_err_coef_"):
            msg = "Summary statistics were not calculated on .fit(). Options to fix:\n"
            msg += "\t- Re-fit using .fit(X, y, calculate_stats=True)\n"
            msg += "\t- Re-inititialize using LogisticRegression(calculate_stats=True)"
            raise AssertionError(msg)

        data = {
            "Coef.": (self.intercept_.tolist() + self.coef_.tolist()[0]),
            "Std.Err": (self.std_err_intercept_.tolist() + self.std_err_coef_.tolist()),
            "z": (self.z_intercept_.tolist() + self.z_coef_.tolist()[0]),
            "P>|z|": (self.p_val_intercept_.tolist() + self.p_val_coef_.tolist()[0]),
        }
        
        stats = pd.DataFrame(data, index=self.names_)
        stats["[ 0.025"] = stats["Coef."] - 1.96 * stats["Std.Err"]
        stats["0.975 ]"] = stats["Coef."] + 1.96 * stats["Std.Err"]
        
        stats["VIF"] = self.vif
        
        return stats
    
    @staticmethod
    def convert_sparse_matrix(x):
        """
        Converts a sparse matrix to a numpy array.
        This can prevent problems arising from, e.g. OneHotEncoder.
        Args:
            x: numpy array, sparse matrix
        Returns:
            numpy array of x
        """
        if scipy.sparse.issparse(x):
            return x.toarray()
        else:
            return x
    
    def plot_weights(self, save=None, figsize=(15, 8), fontsize=14, color=["#2639E9", "#F76E6C", "#FE7715"]):
        summary = self.summary()
        
        x = summary["Coef."]
        y = summary.index
        lower_error = summary["Coef."] - summary["[ 0.025"]
        upper_error = summary["0.975 ]"] - summary["Coef."]
        
        fig, ax = plt.subplots(1, 1, figsize=figsize)
        ax.errorbar(x, y, xerr=[lower_error, upper_error], fmt="o", ecolor=color[0], elinewidth=2, capthick=2, capsize=4, ms=6, mfc=color[0], mec=color[0])
        # ax.tick_params(axis='x', labelrotation=0, grid_color="#FFFFFF", labelsize=fontsize)
        # ax.tick_params(axis='y', labelrotation=0, grid_color="#FFFFFF", labelsize=fontsize)
        ax.axvline(0, color=color[0], linestyle='--', ymax=len(y), alpha=0.5)
        ax.spines['top'].set_color(color[0])
        ax.spines['bottom'].set_color(color[0])
        ax.spines['right'].set_color(color[0])
        ax.spines['left'].set_color(color[0])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        ax.set_title("Regression Meta Analysis - Weight Plot", fontsize=fontsize, fontweight="bold")
        ax.set_xlabel("Weight Estimates", fontsize=fontsize, weight="bold")
        ax.set_ylabel("Variable", fontsize=fontsize, weight="bold")
        
        if save:
            if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
                os.makedirs(os.path.dirname(save))
            
            plt.savefig(save, dpi=240, format="png", bbox_inches="tight")

        return fig

    # def plot_weights(self, save=None):
    #     """
    #     Generates a weight plot(plotly chart) from `stats`
    #     Example:
    #     ```
    #     pipeline = Pipeline([
    #         ('clf', LogisticRegression(calculate_stats=True))
    #     ])
    #     pipeline.fit(X, y)
    #     stats = pipeline.named_steps['clf'].plot_weights()
    #     ```
    #     Args:
    #         stats: The statistics to display
    #         format: The format of the image, such as 'png'. The default None returns a plotly image.
    #         scale: If format is specified, the scale of the image
    #         width: If format is specified, the width of the image
    #         height: If format is specified, the image of the image
    #     """
    #     stats = self.summary()
        
    #     fig = go.Figure()

    #     fig.add_trace(
    #         go.Scatter(
    #             x=stats['Coef.'],
    #             y=stats['Coef.'].index,
    #             line=dict(color='#2639E9', width=2),
    #             mode='markers',

    #             error_x=dict(
    #                 type='data',
    #                 symmetric=False,
    #                 array=stats['0.975 ]'] - stats['Coef.'],
    #                 arrayminus=stats['Coef.'] - stats['[ 0.025'],
    #                 color='#2639E9')
    #         )
    #     )

    #     fig.add_shape(type="line",
    #                 x0=0, y0=0, x1=0, y1=len(stats),
    #                 line=dict(color="#a29bfe", width=3, dash='dash')
    #                 )

    #     fig.update_layout(
    #         title='Regression Meta Analysis - Weight Plot',
    #         xaxis_title='Weight Estimates',
    #         yaxis_title='Variable',
    #         xaxis_showgrid=False,
    #         yaxis_showgrid=False
    #     )
        
    #     fig.update_layout(template="simple_white")
        
    #     if save:
    #         write_image(fig, save)

    #     return fig
    
    
class ScoreCard(toad.ScoreCard, TransformerMixin):
    
    def __init__(self, target="target", pdo=60, rate=2, base_odds=35, base_score=750, combiner={}, transer=None, pretrain_lr=None, pipeline=None, **kwargs):
        """
        评分卡模型转换

        Args:
            target: 数据集中标签名称，默认 target
            pdo: odds 每增加 rate 倍时减少 pdo 分，默认 60
            rate: 倍率
            base_odds: 基础 odds，通常根据业务经验设置的基础比率（违约概率/正常概率），估算方法：（1-样本坏客户占比）/坏客户占比，默认 35，即 35:1 => 0.972 => 坏样本率 2.8%
            base_score: 基础 odds 对应的分数，默认 750
            combiner: 分箱转换器，传入 pipeline 时可以为None
            transer: woe转换器，传入 pipeline 时可以为None
            pretrain_lr: 预训练好的逻辑回归模型，可以不传
            pipeline: 训练好的 pipeline，必须包含 Combiner 和 WOETransformer
            **kwargs: 其他相关参数，具体参考 toad.ScoreCard
        """
        if pipeline:
            combiner = self.class_steps(pipeline, Combiner)[0]
            transer = self.class_steps(pipeline, WOETransformer)[0]
            
            if self.class_steps(pipeline, (ITLubberLogisticRegression, LogisticRegression)):
                pretrain_lr = self.class_steps(pipeline, (ITLubberLogisticRegression, LogisticRegression))[0]
            
        super().__init__(
                            combiner=combiner.combiner if isinstance(combiner, Combiner) else combiner, transer=transer.transformer if isinstance(transer, WOETransformer) else transer, 
                            pdo=pdo, rate=rate, base_odds=base_odds, base_score=base_score, **kwargs
                        )
        
        self.target = target
        self.pipeline = pipeline
        self.pretrain_lr = pretrain_lr
        
    def fit(self, x):
        y = x[self.target]
        x = x.drop(columns=[self.target])
        
        self._feature_names = x.columns.tolist()

        for f in self.features_:
            if f not in self.transer:
                raise Exception('column \'{f}\' is not in transer'.format(f = f))

        if self.pretrain_lr:
            self.model = self.pretrain_lr
        else:
            self.model.fit(x, y)
        
        self.rules = self._generate_rules()

        sub_score = self.woe_to_score(x)
        self.base_effect = pd.Series(np.median(sub_score, axis=0), index = self.features_)

        return self
    
    def transform(self, x):
        return self.predict(x)
    
    @staticmethod
    def KS_bucket(y_pred, y_true, bucket=10, method="quantile"):
        return toad.metrics.KS_bucket(y_pred, y_true, bucket=bucket, method=method)
    
    @staticmethod
    def KS(y_pred, y_true):
        return toad.metrics.KS(y_pred, y_true)
    
    @staticmethod
    def AUC(y_pred, y_true):
        return toad.metrics.AUC(y_pred, y_true)
    
    @staticmethod
    def perf_eva(y_pred, y_true, title="", plot_type=["ks", "roc"], save=None, figsize=(14, 6)):
        # plt.figure(figsize=figsize)
        rt = sc.perf_eva(y_true, y_pred, title=title, plot_type=plot_type, show_plot=True)

        if save:
            if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
                os.makedirs(os.path.dirname(save))
            
            rt["pic"].savefig(save, dpi=240, format="png", bbox_inches="tight")
        
        return rt
    
    @staticmethod
    def ks_plot(score, target, title="", fontsize=14, figsize=(16, 8), save=None, colors=["#2639E9", "#F76E6C", "#FE7715"]):
        if np.mean(score) < 0 or np.mean(score) > 1:
            warnings.warn('Since the average of pred is not in [0,1], it is treated as predicted score but not probability.')
            score = -score

        df = pd.DataFrame({'label': target, 'pred': score})
        def n0(x): return sum(x==0)
        def n1(x): return sum(x==1)
        df_ks = df.sort_values('pred', ascending=False).reset_index(drop=True) \
            .assign(group=lambda x: np.ceil((x.index+1)/(len(x.index)/len(df.index)))) \
            .groupby('group')['label'].agg([n0, n1]) \
            .reset_index().rename(columns={'n0':'good','n1':'bad'}) \
            .assign(
                group=lambda x: (x.index+1)/len(x.index),
                cumgood=lambda x: np.cumsum(x.good)/sum(x.good), 
                cumbad=lambda x: np.cumsum(x.bad)/sum(x.bad)
            ).assign(ks=lambda x:abs(x.cumbad-x.cumgood))

        fig, ax = plt.subplots(1, 2, figsize = figsize)

        # KS曲线
        dfks = df_ks.loc[lambda x: x.ks==max(x.ks)].sort_values('group').iloc[0]

        ax[0].plot(df_ks.group, df_ks.ks, color=colors[0], label="KS曲线")
        ax[0].plot(df_ks.group, df_ks.cumgood, color=colors[1], label="累积好客户占比")
        ax[0].plot(df_ks.group, df_ks.cumbad, color=colors[2], label="累积坏客户占比")
        ax[0].fill_between(df_ks.group, df_ks.cumbad, df_ks.cumgood, color=colors[0], alpha=0.25)

        ax[0].plot([dfks['group'], dfks['group']], [0, dfks['ks']], 'r--')
        ax[0].text(dfks['group'], dfks['ks'], f"KS: {round(dfks['ks'],4)} at: {dfks.group:.2%}", horizontalalignment='center', fontsize=fontsize)

        ax[0].spines['top'].set_color(colors[0])
        ax[0].spines['bottom'].set_color(colors[0])
        ax[0].spines['right'].set_color(colors[0])
        ax[0].spines['left'].set_color(colors[0])
        ax[0].set_xlabel('% of Population', fontsize=fontsize)
        ax[0].set_ylabel('% of Total Bad / Good', fontsize=fontsize)

        ax[0].set_xlim((0, 1))
        ax[0].set_ylim((0, 1))
        
        handles1, labels1 = ax[0].get_legend_handles_labels()

        ax[0].legend(loc='upper center', ncol=len(labels1), bbox_to_anchor=(0.5, 1.1), frameon=False)

        # ROC 曲线
        fpr, tpr, thresholds = roc_curve(target, score)
        auc_value = toad.metrics.AUC(score, target)

        ax[1].plot(fpr, tpr, color=colors[0], label="ROC Curve")
        ax[1].stackplot(fpr, tpr, color=colors[0], alpha=0.25)
        ax[1].plot([0, 1], [0, 1], color=colors[1], lw=2, linestyle=':')
        # ax[1].tick_params(axis='x', labelrotation=0, grid_color="#FFFFFF", labelsize=fontsize)
        # ax[1].tick_params(axis='y', labelrotation=0, grid_color="#FFFFFF", labelsize=fontsize)
        ax[1].text(0.5, 0.5, f"AUC: {auc_value:.4f}", fontsize=fontsize, horizontalalignment="center", transform=ax[1].transAxes)

        ax[1].spines['top'].set_color(colors[0])
        ax[1].spines['bottom'].set_color(colors[0])
        ax[1].spines['right'].set_color(colors[0])
        ax[1].spines['left'].set_color(colors[0])
        ax[1].set_xlabel("False Positive Rate", fontsize=fontsize)
        ax[1].set_ylabel('True Positive Rate', fontsize=fontsize)

        ax[1].set_xlim((0, 1))
        ax[1].set_ylim((0, 1))

        ax[1].yaxis.tick_right()
        ax[1].yaxis.set_label_position("right")

        handles2, labels2 = ax[1].get_legend_handles_labels()

        ax[1].legend(loc='upper center', ncol=len(labels2), bbox_to_anchor=(0.5, 1.1), frameon=False)
        
        if title: title += " "
        fig.suptitle(f"{title}K-S & ROC CURVE\n", fontsize=fontsize, fontweight="bold")
        
        plt.tight_layout()
        
        if save:
            if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
                os.makedirs(os.path.dirname(save))
                
            plt.savefig(save, dpi=240, format="png", bbox_inches="tight")

        return fig
    
    @staticmethod
    def PSI(y_pred_train, y_pred_oot):
        return toad.metrics.PSI(y_pred_train, y_pred_oot)
    
    @staticmethod
    def perf_psi(y_pred_train, y_pred_oot, y_true_train, y_true_oot, keys=["train", "test"], x_limits=None, x_tick_break=50, show_plot=True, return_distr_dat=False):
        return sc.perf_psi(
            score = {keys[0]: y_pred_train, keys[1]: y_pred_oot},
            label = {keys[0]: y_true_train, keys[1]: y_true_oot},
            x_limits = x_limits,
            x_tick_break = x_tick_break,
            show_plot = show_plot,
            return_distr_dat = return_distr_dat,
        )
    
    @staticmethod
    def score_hist(score, y_true, figsize=(15, 10), bins=20, alpha=1, save=None):
        fig, ax = plt.subplots(1, 1, figsize = figsize)
        palette = sns.diverging_palette(340, 267, n=2, s=100, l=40)

        sns.histplot(
                    x=score, hue=y_true.replace({0: "good", 1: "bad"}), element="step", stat="density", bins=bins, common_bins=True, common_norm=True, palette=palette, ax=ax
                )

        sns.despine()

        ax.spines['top'].set_color("#2639E9")
        ax.spines['bottom'].set_color("#2639E9")
        ax.spines['right'].set_color("#2639E9")
        ax.spines['left'].set_color("#2639E9")

        ax.set_xlabel("score")
        ax.set_ylabel("density")
        
        ax.legend(["坏样本", "好样本"], loc='upper center', ncol=len(y_true.unique()), bbox_to_anchor=(0.5, 1.05), frameon=False, fontsize=14)
        
        fig.tight_layout()

        if save:
            if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
                os.makedirs(os.path.dirname(save))
                
            plt.savefig(save, dpi=240, format="png", bbox_inches="tight")
        
        return fig
    
    def _format_rule(self, rule, decimal = 2, **kwargs):
        bins = self.format_bins(rule['bins'])
        scores = np.around(rule['scores'], decimals = decimal).tolist()
        
        return dict(zip(bins, scores))
    
    @staticmethod
    def class_steps(pipeline, query):
        return [v for k, v in pipeline.named_steps.items() if isinstance(v, query)]
    
    @staticmethod
    def feature_bins(bins):
        if isinstance(bins, list): bins = np.array(bins)
        EMPTYBINS = len(bins) if not isinstance(bins[0], (set, list, np.ndarray)) else -1
        
        l = []
        if np.issubdtype(bins.dtype, np.number):
            has_empty = len(bins) > 0 and np.isnan(bins[-1])
            if has_empty: bins = bins[:-1]
            sp_l = ["负无穷"] + bins.tolist() + ["正无穷"]
            for i in range(len(sp_l) - 1): l.append('['+str(sp_l[i])+' , '+str(sp_l[i+1])+')')
            if has_empty: l.append('缺失值')
        else:
            for keys in bins:
                keys_update = set()
                for key in keys:
                    if pd.isnull(key) or key == "nan":
                        keys_update.add("缺失值")
                    elif key.strip() == "":
                        keys_update.add("空字符串")
                    else:
                        keys_update.add(key)
                label = ','.join(keys_update)
                l.append(label)

        return {i if b != "缺失值" else EMPTYBINS: b for i, b in enumerate(l)}
    
    def feature_bin_stats(self, data, feature, target="target", rules={}, empty_separate=True, method='step', max_n_bins=10, clip_v=None, desc="评分卡分数", verbose=0, combiner=None, ks=False):
        if method not in ['dt', 'chi', 'quantile', 'step', 'kmeans', 'cart']:
            raise "method is the one of ['dt', 'chi', 'quantile', 'step', 'kmeans', 'cart']"
        
        if combiner is None:
            combiner = toad.transform.Combiner()
            
            if method == "cart":
                x = data[feature].values
                y = data[target]
                _combiner = OptimalBinning(feature, dtype="numerical", max_n_bins=max_n_bins, monotonic_trend="auto_asc_desc", gamma=0.01).fit(x, y)
                if _combiner.status == "OPTIMAL":
                    rules.update({feature: [s.tolist() if isinstance(s, np.ndarray) else s for s in _combiner.splits] + [np.nan]})
            else:
                combiner.fit(data[[feature, target]], target, empty_separate=empty_separate, method=method, n_bins=max_n_bins, clip_v=clip_v)

            if verbose > 0:
                print(data[feature].describe())

        if rules and isinstance(rules, list): rules = {feature: rules}
        if rules and isinstance(rules, dict): combiner.update(rules)

        feature_bin = combiner.export()[feature]
        feature_bin_dict = self.feature_bins(np.array(feature_bin))
        
        df_bin = combiner.transform(data[[feature, target]], labels=False)
        
        table = df_bin[[feature, target]].groupby([feature, target]).agg(len).unstack()
        table.columns.name = None
        table = table.rename(columns = {0 : '好样本数', 1 : '坏样本数'}).fillna(0)
        if "好样本数" not in table.columns:
            table["好样本数"] = 0
        if "坏样本数" not in table.columns:
            table["坏样本数"] = 0
        
        table["指标名称"] = feature
        table["指标含义"] = desc
        table = table.reset_index().rename(columns={feature: "分箱"})

        table['样本总数'] = table['好样本数'] + table['坏样本数']
        table['样本占比'] = table['样本总数'] / table['样本总数'].sum()
        table['好样本占比'] = table['好样本数'] / table['好样本数'].sum()
        table['坏样本占比'] = table['坏样本数'] / table['坏样本数'].sum()
        table['坏样本率'] = table['坏样本数'] / table['样本总数']
        
        table = table.fillna(0.)
        
        table['分档WOE值'] = table.apply(lambda x : np.log(x['好样本占比'] / (x['坏样本占比'] + 1e-6)),axis=1)
        table['分档IV值'] = table.apply(lambda x : (x['好样本占比'] - x['坏样本占比']) * np.log(x['好样本占比'] / (x['坏样本占比'] + 1e-6)), axis=1)
        
        table = table.replace(np.inf, 0).replace(-np.inf, 0)
        
        table['指标IV值'] = table['分档IV值'].sum()
        
        table["LIFT值"] = table['坏样本率'] / (table["坏样本数"].sum() / table["样本总数"].sum())
        table["累积LIFT值"] = (table['坏样本数'].cumsum() / table['样本总数'].cumsum()) / (table["坏样本数"].sum() / table["样本总数"].sum())
        # table["累积LIFT值"] = table["LIFT值"].cumsum()
        
        if ks:
            table = table.sort_values("分箱")
            table["累积好样本数"] = table["好样本数"].cumsum()
            table["累积坏样本数"] = table["坏样本数"].cumsum()
            table["分档KS值"] = table["累积坏样本数"] / table['坏样本数'].sum() - table["累积好样本数"] / table['好样本数'].sum()
        
        table["分箱"] = table["分箱"].map(feature_bin_dict)
        
        if ks:
            return table[['指标名称', "指标含义", '分箱', '样本总数', '样本占比', '好样本数', '好样本占比', '坏样本数', '坏样本占比', '坏样本率', '分档WOE值', '分档IV值', '指标IV值', 'LIFT值', '累积LIFT值', '累积好样本数', '累积坏样本数', '分档KS值']]
        else:
            return table[['指标名称', "指标含义", '分箱', '样本总数', '样本占比', '好样本数', '好样本占比', '坏样本数', '坏样本占比', '坏样本率', '分档WOE值', '分档IV值', '指标IV值', 'LIFT值', '累积LIFT值']]

    
if __name__ == '__main__':
    # https://github.com/itlubber/openpyxl-excel-style-template/blob/main/pipeline_model.py
    plt.ion()
    
    target = "creditability"
    data = sc.germancredit()
    data[target] = data[target].map({"good": 0, "bad": 1})

    train, test = train_test_split(data, test_size=0.3, shuffle=True, stratify=data[target])
    oot = data.copy()
    feature_pipeline = Pipeline([
        ("preprocessing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("combiner", Combiner(target=target, min_samples=0.2)),
        ("transform", WOETransformer(target=target)),
        ("processing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("stepwise", StepwiseSelection(target=target)),
    ])

    feature_pipeline.fit(train)

    woe_train = feature_pipeline.transform(train)
    woe_test = feature_pipeline.transform(test)
    woe_oot = feature_pipeline.transform(oot)
    
    # save all bin_plot
    _combiner = feature_pipeline.named_steps["combiner"]
    for col in woe_train.columns:
        if col != target:
            _combiner.bin_plot(train, col, labels=True, save=f"outputs/bin_plots/train_{col}.png")
            _combiner.bin_plot(test, col, labels=True, save=f"outputs/bin_plots/test_{col}.png")
            _combiner.bin_plot(oot, col, labels=True, save=f"outputs/bin_plots/oot_{col}.png")

    # logistic = StatsLogisticRegression(target=target)
    logistic = ITLubberLogisticRegression(target=target)
    
    logistic.fit(woe_train)

    y_pred_train = logistic.predict_proba(woe_train.drop(columns=target))[:, 1]
    y_pred_test = logistic.predict_proba(woe_test.drop(columns=target))[:, 1]
    y_pred_oot = logistic.predict_proba(woe_oot.drop(columns=target))[:, 1]
    
    # params_grid = {
    #     # "logistic__C": [i / 1. for i in range(1, 10, 2)],
    #     # "logistic__penalty": ["l2"],
    #     # "logistic__class_weight": [None, "balanced"], # + [{1: i / 10.0, 0: 1 - i / 10.0} for i in range(1, 10)],
    #     # "logistic__max_iter": [100],
    #     # "logistic__solver": ["sag"] # ["liblinear", "sag", "lbfgs", "newton-cg"],
    #     "logistic__intercept": [True, False],
    # }
    
    # clf = GridSearchCV(feature_pipeline, params_grid, cv=5, scoring='roc_auc', verbose=-1, n_jobs=2, return_train_score=True)
    # clf.fit(train, train[target])
    
    # y_pred_train = clf.best_estimator_.predict(train)
    # y_pred_test = clf.best_estimator_.predict(test)
    
    # print(clf.best_params_)
    
    # model summary
    # logistic.summary_save()
    
    logistic.plot_weights(save="outputs/logistic_train.png")
    
    summary = logistic.summary().reset_index().rename(columns={"index": "Features"})
    
    train_corr = logistic.corr(woe_train, save="outputs/train_corr.png")
    test_corr = logistic.corr(woe_test, save="outputs/test_corr.png")
    oot_corr = logistic.corr(woe_oot, save="outputs/oot_corr.png")
    
    train_report = logistic.report(woe_train)
    test_report = logistic.report(woe_test)
    oot_report = logistic.report(woe_oot)
    
    print("train: ", toad.metrics.KS(y_pred_train, train[target]), toad.metrics.AUC(y_pred_train, train[target]))
    print("test: ", toad.metrics.KS(y_pred_test, test[target]), toad.metrics.AUC(y_pred_test, test[target]))
    print("oot: ", toad.metrics.KS(y_pred_oot, oot[target]), toad.metrics.AUC(y_pred_oot, oot[target]))

    card = ScoreCard(target=target, pipeline=feature_pipeline, pretrain_lr=logistic)
    card.fit(woe_train)
    
    train["score"] = card.predict(train)
    test["score"] = card.predict(test)
    oot["score"] = card.predict(oot)
    
    card.perf_eva(train["score"], train[target], title="Train Dataset", save="outputs/train_ksplot.png")
    card.perf_eva(test["score"], test[target], title="Test Dataset", save="outputs/test_ksplot.png")
    card.perf_eva(oot["score"], oot[target], title="OOT Dataset", save="outputs/oot_ksplot.png")
    
    card.score_hist(train["score"], train[target], save="outputs/train_scorehist.png")
    card.score_hist(test["score"], test[target], save="outputs/test_scorehist.png")
    card.score_hist(oot["score"], oot[target], save="outputs/oot_scorehist.png")
    
    train_score_rank = card.feature_bin_stats(train, "score", target=target, rules=[i for i in range(400, 800, 50)], verbose=0, method="step")
    test_score_rank = card.feature_bin_stats(test, "score", target=target, rules=[i for i in range(400, 800, 50)], verbose=0, method="step")
    oot_score_rank = card.feature_bin_stats(oot, "score", target=target, rules=[i for i in range(400, 800, 50)], verbose=0, method="step")
    
    card_points = card.export(to_frame=True)
    
    writer = pd.ExcelWriter("outputs/评分卡结果验证表.xlsx", engine="openpyxl")
    
    summary.to_excel(writer, sheet_name="逻辑回归拟合结果", startrow=1, index=False)
    train_report.to_excel(writer, sheet_name="逻辑回归拟合结果", startrow=len(summary) + 5, index=False)
    test_report.to_excel(writer, sheet_name="逻辑回归拟合结果", startrow=len(summary) + len(train_report) + 9, index=False)
    oot_report.to_excel(writer, sheet_name="逻辑回归拟合结果", startrow=len(summary) + len(train_report) + len(test_report) + 13, index=False)
    
    worksheet = writer.sheets['逻辑回归拟合结果']
    worksheet.cell(row=1, column=1).value = "入模变量系数及相关统计指标"
    worksheet.cell(row=len(summary) + 5, column=1).value = "训练数据集模型预测报告"
    worksheet.cell(row=len(summary) + len(train_report) + 9, column=1).value = "测试数据集模型预测报告"
    worksheet.cell(row=len(summary) + len(train_report) + len(test_report) + 13, column=1).value = "跨时间验证集模型预测报告"
    
    train_corr.to_excel(writer, sheet_name="入模变量相关性", startrow=1, index=True)
    test_corr.to_excel(writer, sheet_name="入模变量相关性", startrow=len(train_corr) + 5, index=True)
    oot_corr.to_excel(writer, sheet_name="入模变量相关性", startrow=len(train_corr) + len(test_corr) + 9, index=True)
    
    worksheet = writer.sheets['入模变量相关性']
    worksheet.cell(row=2, column=1).value = "训练数据集入模变量相关性"
    worksheet.cell(row=len(train_corr) + 6, column=1).value = "测试数据集入模变量相关性"
    worksheet.cell(row=len(train_corr) + len(test_corr) + 10, column=1).value = "跨时间验证集入模变量相关性"
    
    card_points.to_excel(writer, sheet_name="评分卡", index=False)
    
    train_score_rank.to_excel(writer, sheet_name="评分卡排序性", startrow=1, index=False)
    test_score_rank.to_excel(writer, sheet_name="评分卡排序性", startrow=len(train_score_rank) + 5, index=False)
    oot_score_rank.to_excel(writer, sheet_name="评分卡排序性", startrow=len(train_score_rank) + len(test_score_rank) + 9, index=False)
    
    worksheet = writer.sheets['评分卡排序性']
    
    worksheet.cell(row=1, column=1).value = "训练数据集评分排序性"
    worksheet.cell(row=len(train_score_rank) + 5, column=1).value = "测试数据集评分排序性"
    worksheet.cell(row=len(train_score_rank) + len(test_score_rank) + 9, column=1).value = "跨时间验证集评分排序性"
    
    writer.close()
    
    from utils.tools import render_excel
    
    render_excel("outputs/评分卡结果验证表.xlsx", border=False)
    