import os
import cairosvg
import graphviz
import dtreeviz
import warnings
import numpy as np
import pandas as pd

import category_encoders as ce
from sklearn.preprocessing import LabelEncoder
from sklearn.tree import _tree, DecisionTreeClassifier, plot_tree, export_graphviz


warnings.filterwarnings("ignore")


class ParseDecisionTreeRules:
    
    def __init__(self, target="target", labels=["positive", "negative"], feature_map={}, nan=-1., max_iter=128, output="model_report/auto_mining_rules/决策树组合策略挖掘.xlsx", writer=None):
        self.target = target
        self.labels = labels
        self.feature_map = feature_map
        self.nan = nan
        self.max_iter = max_iter
        self.output = output
        self.decision_trees = []
        self.target_enc = None
        self.feature_names = None
        self.dt_rules = pd.DataFrame()
        self.end_row = 2
        self.start_col = 2
        
        if output:
            from utils.excel_writer import ExcelWriter
            from openpyxl.utils import get_column_letter, column_index_from_string
            init_setting()
            if writer:
                self.writer = writer
            else:
                self.writer = ExcelWriter(style_excel="./utils/报告输出模版.xlsx", theme_color="2639E9")
            
            self.worksheet = self.writer.get_sheet_by_name("决策树组合策略挖掘")
    
    def encode_cat_features(self, X, y):
        cat_features = list(set(X.select_dtypes(include=[object, pd.CategoricalDtype]).columns))
        cat_features_index = [i for i, f in enumerate(X.columns) if f in cat_features]
        
        if len(cat_features) > 0:
            if self.target_enc is None:
                self.target_enc = ce.TargetEncoder(cols=cat_features)
                self.target_enc.fit(X[cat_features], y)
                self.target_enc.target_mapping = {}
                X_TE = X.join(self.target_enc.transform(X[cat_features]).add_suffix('_target'))
                for col in cat_features:
                    mapping = X_TE[[col, f"{col}_target"]].drop_duplicates()
                    self.target_enc.target_mapping[col] = dict(zip(mapping[col], mapping[f"{col}_target"]))
            else:
                X_TE = X.join(self.target_enc.transform(X[cat_features]).add_suffix('_target'))
            
            X_TE = X_TE.drop(columns=cat_features)
            return X_TE.rename(columns={f"{c}_target": c for c in cat_features})
        else:
            return X
    
    @staticmethod
    def get_dt_rules(tree, feature_names, total_bad_rate, total_count):
        tree_ = tree.tree_
        left = tree.tree_.children_left
        right = tree.tree_.children_right
        feature_name = [feature_names[i] if i != -2 else "undefined!" for i in tree_.feature]
        rules=dict()

        global res_df
        res_df = pd.DataFrame()

        def recurse(node, depth, parent): # 搜每个节点的规则

            if tree_.feature[node] != -2:  # 非叶子节点,搜索每个节点的规则
                name = feature_name[node]
                thd = np.round(tree_.threshold[node],3)
                s= "{} <= {} ".format( name, thd, node )
                # 左子
                if node == 0:
                    rules[node]=s
                else:
                    rules[node]=rules[parent]+' & ' +s
                recurse(left[node], depth + 1, node)
                s="{} > {}".format(name, thd)
                # 右子 
                if node == 0:
                    rules[node]=s
                else:
                    rules[node]=rules[parent]+' & ' +s
                recurse(right[node], depth + 1, node)
            else:
                df = pd.DataFrame()
                df['组合策略'] = rules[parent],
                df['好样本数'] = tree_.value[node][0][0].astype(int)
                df['好样本占比'] = df['好样本数'] / (total_count * (1 - total_bad_rate))
                df['坏样本数'] = tree_.value[node][0][1].astype(int)
                df['坏样本占比'] = df['坏样本数'] / (total_count * total_bad_rate)
                df['命中数'] = df['好样本数'] + df['坏样本数']
                df['命中率'] = df['命中数'] / total_count
                df['坏率'] = df['坏样本数'] / df['命中数']
                df['样本整体坏率'] = total_bad_rate
                df['LIFT值'] = df['坏率'] / df['样本整体坏率']

                global res_df

                res_df = pd.concat([res_df, df], 0)

        recurse(0, 1, 0)

        return res_df.sort_values("LIFT值", ascending=True).reset_index(drop=True)
    
    @staticmethod
    def select_dt_rules(decision_tree, x, y, lift=3., max_samples=0.05, labels=["positive", "negative"], save=None, verbose=False, drop=False):
        rules = self.get_dt_rules(decision_tree, x.columns, sum(y) / len(y), len(y))
        viz_model = dtreeviz.model(decision_tree,
                                   X_train=x, 
                                   y_train=y,
                                   feature_names=x.columns,
                                   target_name=target, 
                                   class_names=labels,
                                  )
        rules = rules.query(f"LIFT值 >= {lift} & 命中率 <= {max_samples}").reset_index(drop=True)

        if len(rules) > 0:
            decision_tree_viz = viz_model.view(
                                                scale=1.5, 
                                                orientation='LR', 
                                                colors={
                                                        "classes": [None, None, ["#2639E9", "#F76E6C"], ["#2639E9", "#F76E6C", "#FE7715", "#FFFFFF"]],
                                                        "arrow": "#2639E9",
                                                        'text_wedge': "#F76E6C",
                                                        "pie": "#2639E9",
                                                        "tile_alpha": 1,
                                                        "legend_edge": "#FFFFFF",
                                                    },
                                                ticks_fontsize=10,
                                                label_fontsize=10,
                                            )
            if verbose:
                display(rules)
                display(decision_tree_viz)
            if save:
                if os.path.dirname(save) and not os.path.exists(os.path.dirname(save)):
                    os.makedirs(os.path.dirname(save))

                decision_tree_viz.save("combine_rules_cache.svg")
                cairosvg.svg2png(url="combine_rules_cache.svg", write_to=save, dpi=240)

        if drop:
            return rules, decision_tree.feature_names_in_[list(decision_tree.feature_importances_).index(max(decision_tree.feature_importances_))]
        else:
            return rules
    
    @staticmethod
    def query_dt_rules(x, y, parsed_rules=None):
        total_count = len(y)
        total_bad_rate = y.sum() / len(y)

        rules = pd.DataFrame()
        for rule in parsed_rules["组合策略"].unique():
            select_index = x.query(rule).index
            if len(select_index) > 0:
                y_select = y[select_index]
                df = pd.Series()
                df['组合策略'] = rule
                df['好样本数'] = len(y_select) - y_select.sum()
                df['好样本占比'] = df['好样本数'] / (total_count * (1 - total_bad_rate))
                df['坏样本数'] = y_select.sum()
                df['坏样本占比'] = df['坏样本数'] / (total_count * total_bad_rate)
                df['命中数'] = df['好样本数'] + df['坏样本数']
                df['命中率'] = df['命中数'] / total_count
                df['坏率'] = df['坏样本数'] / df['命中数']
                df['样本整体坏率'] = total_bad_rate
                df['LIFT值'] = df['坏率'] / df['样本整体坏率']
            else:
                df = pd.Series({'组合策略': rule,'好样本数': 0,'好样本占比': 0.,'坏样本数': 0,'坏样本占比': 0.,'命中数': 0,'命中率': 0.,'坏率': 0.,'样本整体坏率': total_bad_rate,'LIFT值': 0.,})

            rules = pd.concat([rules, pd.DataFrame(df).T]).reset_index(drop=True)

        return rules
    
    def insert_dt_rules(self, parsed_rules, end_row, start_col, save=None):
        end_row, end_col = self.writer.insert_df2sheet(self.worksheet, parsed_rules, (end_row + 2, start_col))
        
        for c in ['好样本占比', '坏样本占比', '命中率', '坏率', '样本整体坏率', 'LIFT值']:
            conditional_column = get_column_letter(start_col + parsed_rules.columns.get_loc(c))
            self.writer.set_number_format(self.worksheet, f"{conditional_column}{end_row - len(parsed_rules)}:{conditional_column}{end_row - 1}", "0.00%")
        for c in ["坏率", "LIFT值"]:
            conditional_column = get_column_letter(start_col + parsed_rules.columns.get_loc(c))
            self.writer.add_conditional_formatting(self.worksheet, f'{conditional_column}{end_row - len(parsed_rules)}', f'{conditional_column}{end_row - 1}')
        
        if save is not None:
            end_row, end_col = self.writer.insert_pic2sheet(self.worksheet, save, (end_row + 1, start_col), figsize=(400, 300))
        
        return end_row, end_col
        
    def fit(self, x, y=None, max_depth=2, lift=3, max_samples=0.2, min_score=None, verbose=False, **kwargs):
        y = x[self.target]
        X_TE = self.encode_cat_features(x.drop(columns=[self.target]), y)
        X_TE = X_TE.fillna(self.nan)
        
        self.feature_names = list(X_TE.columns)
        
        for i in range(self.max_iter):
            decision_tree = DecisionTreeClassifier(max_depth=max_depth, **kwargs)
            decision_tree = decision_tree.fit(X_TE, y)
            
            if min_score is not None and decision_tree.score(X_TE, y) < min_score:
                break
            
            parsed_rules, remove = self.select_dt_rules(decision_tree, X_TE, y, lift=lift, max_samples=max_samples, labels=self.labels, verbose=verbose, save=f"model_report/auto_mining_rules/combiner_rules_{i}.png", drop=True)
            
            if len(parsed_rules) > 0:
                self.dt_rules = pd.concat([self.dt_rules, parsed_rules]).reset_index(drop=True)
                
                if self.writer is not None:
                    parsed_rules["组合策略"] = parsed_rules["组合策略"].replace(self.feature_map, regex=True)
                    self.end_row, _ = self.insert_dt_rules(parsed_rules, self.end_row, self.start_col, save=f"model_report/auto_mining_rules/combiner_rules_{i}.png")
                    
            X_TE = X_TE.drop(columns=remove)
            self.decision_trees.append(decision_tree)
        
        return self
    
    def transform(self, x, y=None):
        y = x[self.target]
        X_TE = self.encode_cat_features(x.drop(columns=[self.target]), y)
        X_TE = X_TE.fillna(self.nan)
        parsed_rules = self.query_dt_rules(X_TE, y, parsed_rules=self.dt_rules)
        parsed_rules["组合策略"] = parsed_rules["组合策略"].replace(self.feature_map, regex=True)
        return parsed_rules
    
    def insert_all_rules(self, val=None, test=None):
        parsed_rules_train = self.dt_rules.copy()
        parsed_rules_train["组合策略"] = parsed_rules_train["组合策略"].replace(self.feature_map, regex=True)
        self.end_row, _ = self.writer.insert_value2sheet(self.worksheet, (self.end_row + 2, self.start_col), value="训练集决策树组合策略")
        self.end_row, _ = self.insert_dt_rules(parsed_rules_train, self.end_row, self.start_col)
        
        if val is not None:
            parsed_rules_val = self.transform(val)
            self.end_row, _ = self.writer.insert_value2sheet(self.worksheet, (self.end_row + 2, self.start_col), value="验证集决策树组合策略")
            self.end_row, _ = self.insert_dt_rules(parsed_rules_val, self.end_row, self.start_col)
        
        if test is not None:
            parsed_rules_test = self.transform(test)
            self.end_row, _ = self.writer.insert_value2sheet(self.worksheet, (self.end_row + 2, self.start_col), value="测试集决策树组合策略")
            self.end_row, _ = self.insert_dt_rules(parsed_rules_test, self.end_row, self.start_col)
            
    def save(self):
        self.writer.save(self.output)
        
       
if __name__ == '__main__':
    pdtr = ParseDecisionTreeRules(target=target, feature_map=feature_map, max_iter=8)
    pdtr.fit(train, lift=3., max_depth=2, max_samples=0.1, verbose=False, min_samples_split=8, min_samples_leaf=5, max_features="auto")
    pdtr.insert_all_rules(test=test)
    pdtr.save()
